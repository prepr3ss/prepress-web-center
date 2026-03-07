# Plan Scraper Routes
from datetime import datetime
from flask import Blueprint, jsonify, request, render_template, current_app, flash, redirect, url_for
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import openpyxl
import os
import pytz
from sqlalchemy import or_, and_, extract, cast, String

# Import blueprint
from . import plan_scraper_bp

# Import functions to avoid circular imports
def get_db():
    from app import db
    return db

def get_require_mounting_access():
    from app import require_mounting_access
    return require_mounting_access

def get_plan_scraper_model():
    from .models import PlanScraperData
    return PlanScraperData

# Import production models - lazy import to avoid blueprint registration issues
def get_production_models():
    from .production_models import (
        ProductionCustomerName,
        ProductionImpositionCockpit,
        ProductionImpositionRemarks,
        ProductionImpositionJob,
        ProductionPrintMachine
    )
    return {
        'ProductionCustomerName': ProductionCustomerName,
        'ProductionImpositionCockpit': ProductionImpositionCockpit,
        'ProductionImpositionRemarks': ProductionImpositionRemarks,
        'ProductionImpositionJob': ProductionImpositionJob,
        'ProductionPrintMachine': ProductionPrintMachine
    }

def get_work_queue_model():
    from .models import WorkQueue
    return WorkQueue

def get_production_imposition_job_model():
    from .production_models import ProductionImpositionJob
    return ProductionImpositionJob

# Timezone untuk Jakarta
jakarta_tz = pytz.timezone('Asia/Jakarta')

# Allowed file extensions
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    """Check if file has allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_excel_file(file_path):
    """Process Excel file and extract data from specific sheets"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Sheets to process (exclude DP sheet)
        target_sheets = ['SM2', 'SM3', 'SM4', 'SM5', 'SM6', 'VLF']
        extracted_data = []
        
        print(f"Available sheets in workbook: {workbook.sheetnames}")
        
        for sheet_name in target_sheets:
            actual_sheet_name = None
            
            # Try exact match first
            if sheet_name in workbook.sheetnames:
                actual_sheet_name = sheet_name
            else:
                # Try case-insensitive match
                for available_sheet in workbook.sheetnames:
                    if available_sheet.strip().upper() == sheet_name.upper():
                        actual_sheet_name = available_sheet
                        break
                
                # Try with spaces (e.g., 'SM 2' instead of 'SM2')
                if actual_sheet_name is None:
                    spaced_variants = [
                        sheet_name.replace('SM', 'SM '),
                        sheet_name.replace('SM', ' SM'),
                        sheet_name + ' ',
                        ' ' + sheet_name
                    ]
                    for variant in spaced_variants:
                        if variant in workbook.sheetnames:
                            actual_sheet_name = variant
                            break
            
            if actual_sheet_name is None:
                print(f"Sheet '{sheet_name}' not found in workbook")
                continue
                
            sheet = workbook[actual_sheet_name]
            print(f"Processing sheet: {actual_sheet_name} (target: {sheet_name})")
            
            # Find header row with column names
            header_row = None
            column_mapping = {}
            
            for row in sheet.iter_rows(min_row=1, max_row=10):  # Check first 10 rows for headers
                row_values = [str(cell.value).strip() if cell.value else '' for cell in row]
                
                # Check for required headers (case-insensitive)
                has_wo_sap = any('WO SAP' in val.upper() for val in row_values)
                has_no_mc_sap = any('NO MC SAP' in val.upper() for val in row_values)
                
                if has_wo_sap and has_no_mc_sap:
                    header_row = row[0].row
                    print(f"Found header row at: {header_row}")
                    
                    # Create column mapping (case-insensitive, but more specific)
                    for idx, value in enumerate(row_values):
                        val_upper = value.upper()
                        val_exact = value.strip()
                        
                        # Exact matches first (case-insensitive)
                        if val_exact.upper() == 'WO SAP':
                            column_mapping['wo_number'] = idx
                            print(f"Found WO SAP at column {idx}: '{value}'")
                        elif val_exact.upper() == 'NO MC SAP':
                            column_mapping['mc_number'] = idx
                            print(f"Found NO MC SAP at column {idx}: '{value}'")
                        elif val_exact.upper() == 'JENIS BARANG':
                            column_mapping['item_name'] = idx
                            print(f"Found Jenis Barang at column {idx}: '{value}'")
                        elif val_exact.upper() == 'UP':
                            column_mapping['num_up'] = idx
                            print(f"Found Up at column {idx}: '{value}'")
                        elif val_exact.upper() == 'SHEET':
                            column_mapping['run_length_sheet'] = idx
                            print(f"Found Sheet at column {idx}: '{value}'")
                        elif val_exact.upper() == 'KETERANGAN':
                            column_mapping['paper_desc'] = idx
                            print(f"Found Keterangan at column {idx}: '{value}'")
                        elif val_exact.upper() in ['SUPLAYER KERTAS', 'SUPPLIER KERTAS']:
                            column_mapping['paper_type'] = idx
                            print(f"Found Suplayer kertas at column {idx}: '{value}'")
                        # Fallback to partial matches only if exact matches don't work
                        elif 'WO SAP' in val_upper and 'wo_number' not in column_mapping:
                            column_mapping['wo_number'] = idx
                            print(f"Found WO SAP (partial) at column {idx}: '{value}'")
                        elif 'NO MC SAP' in val_upper and 'mc_number' not in column_mapping:
                            column_mapping['mc_number'] = idx
                            print(f"Found NO MC SAP (partial) at column {idx}: '{value}'")
                        elif 'JENIS BARANG' in val_upper and 'item_name' not in column_mapping:
                            column_mapping['item_name'] = idx
                            print(f"Found Jenis Barang (partial) at column {idx}: '{value}'")
                        elif val_upper == 'UP' and 'num_up' not in column_mapping:
                            column_mapping['num_up'] = idx
                            print(f"Found Up (exact) at column {idx}: '{value}'")
                        elif val_upper == 'SHEET' and 'run_length_sheet' not in column_mapping:
                            column_mapping['run_length_sheet'] = idx
                            print(f"Found Sheet (exact) at column {idx}: '{value}'")
                        elif 'KETERANGAN' in val_upper and 'paper_desc' not in column_mapping:
                            column_mapping['paper_desc'] = idx
                            print(f"Found Keterangan (partial) at column {idx}: '{value}'")
                        elif ('SUPLAYER KERTAS' in val_upper or 'SUPPLIER KERTAS' in val_upper) and 'paper_type' not in column_mapping:
                            column_mapping['paper_type'] = idx
                            print(f"Found Suplayer kertas (partial) at column {idx}: '{value}'")
                    
                    print(f"Column mapping: {column_mapping}")
                    break
            
            if header_row and column_mapping:
                # Extract data rows
                for row in sheet.iter_rows(min_row=header_row + 1):
                    row_values = [cell.value for cell in row]
                    
                    # Skip empty rows
                    if not any(row_values):
                        continue
                    
                    # Extract data using column mapping
                    try:
                        # Helper function to safely extract and convert values
                        def get_cell_value(col_idx):
                            if col_idx is None or col_idx >= len(row_values):
                                return None
                            cell_value = row_values[col_idx]
                            return cell_value
                        
                        # Extract raw values
                        wo_raw = get_cell_value(column_mapping.get('wo_number'))
                        mc_raw = get_cell_value(column_mapping.get('mc_number'))
                        item_raw = get_cell_value(column_mapping.get('item_name'))
                        up_raw = get_cell_value(column_mapping.get('num_up'))
                        sheet_raw = get_cell_value(column_mapping.get('run_length_sheet'))
                        desc_raw = get_cell_value(column_mapping.get('paper_desc'))
                        type_raw = get_cell_value(column_mapping.get('paper_type'))
                        
                        # Convert to appropriate types with better error handling
                        wo_number = str(wo_raw).strip() if wo_raw is not None and str(wo_raw).strip() else ''
                        mc_number = str(mc_raw).strip() if mc_raw is not None and str(mc_raw).strip() else ''
                        item_name = str(item_raw).strip() if item_raw is not None and str(item_raw).strip() else ''
                        
                        # Handle numeric values more robustly
                        try:
                            if up_raw is not None:
                                num_up = int(float(str(up_raw).replace(',', '.'))) if str(up_raw).strip() else 0
                            else:
                                num_up = 0
                        except (ValueError, TypeError):
                            num_up = 0
                        
                        try:
                            if sheet_raw is not None:
                                # Handle both integer and string representations
                                sheet_str = str(sheet_raw).strip().replace(',', '.')
                                run_length_sheet = float(sheet_str) if sheet_str else 0.0
                            else:
                                run_length_sheet = 0.0
                        except (ValueError, TypeError):
                            run_length_sheet = 0.0
                        
                        paper_desc = str(desc_raw).strip() if desc_raw is not None and str(desc_raw).strip() else ''
                        paper_type = str(type_raw).strip() if type_raw is not None and str(type_raw).strip() else ''
                        
                        # Skip if essential data is missing
                        if not wo_number or not mc_number or not item_name:
                            continue
                        
                        # Use clean machine name without spaces for storage
                        clean_machine_name = actual_sheet_name.replace(' ', '').strip()
                        extracted_data.append({
                            'print_machine': clean_machine_name,
                            'wo_number': wo_number,
                            'mc_number': mc_number,
                            'item_name': item_name,
                            'num_up': num_up,
                            'run_length_sheet': run_length_sheet,
                            'paper_desc': paper_desc,
                            'paper_type': paper_type
                        })
                        
                    except (ValueError, TypeError, IndexError) as e:
                        # Skip rows with invalid data
                        continue
        
        # Aggregate data by WO number
        PlanScraperData = get_plan_scraper_model()
        aggregated_data = PlanScraperData.aggregate_sheet_by_wo(extracted_data)
        
        # Count successfully processed sheets
        processed_sheets = []
        for sheet_name in target_sheets:
            if sheet_name in workbook.sheetnames:
                processed_sheets.append(sheet_name)
            else:
                # Check case-insensitive
                for available_sheet in workbook.sheetnames:
                    if available_sheet.strip().upper() == sheet_name.upper():
                        processed_sheets.append(sheet_name)
                        break
        
        return {
            'success': True,
            'data': aggregated_data,
            'total_records': len(aggregated_data),
            'message': f'Successfully processed {len(aggregated_data)} records from {len(processed_sheets)} sheets'
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Error processing Excel file: {str(e)}'
        }

@plan_scraper_bp.route('/plan-scraper')
@login_required
def plan_scraper_page():
    """Render halaman Plan Scraper"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(lambda: render_template('plan_scraper/dashboard.html'))()

@plan_scraper_bp.route('/plan-scraper/upload', methods=['POST'])
@login_required
def upload_excel_file():
    """Handle Excel file upload and processing"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_upload_excel_file_impl)()

def _upload_excel_file_impl():
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Invalid file type. Only Excel files (.xlsx, .xls) are allowed'}), 400
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        upload_folder = current_app.config.get('UPLOADS_PATH', 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)
        
        file_path = os.path.join(upload_folder, f"plan_scraper_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
        file.save(file_path)
        
        # Process the Excel file
        result = process_excel_file(file_path)
        
        if result['success']:
            # Save processed data to database
            db = get_db()
            PlanScraperData = get_plan_scraper_model()
            
            saved_count = 0
            updated_count = 0
            
            for data_item in result['data']:
                existing_record = PlanScraperData.query.filter_by(wo_number=data_item['wo_number']).first()
                
                if existing_record:
                    # Update existing record
                    for key, value in data_item.items():
                        if hasattr(existing_record, key):
                            setattr(existing_record, key, value)
                    existing_record.updated_at = datetime.now(jakarta_tz)
                    updated_count += 1
                else:
                    # Create new record
                    new_record = PlanScraperData(
                        created_by=current_user.id,
                        **data_item
                    )
                    db.session.add(new_record)
                    saved_count += 1
            
            db.session.commit()
            
            # Clean up uploaded file
            try:
                os.remove(file_path)
            except:
                pass
            
            return jsonify({
                'success': True,
                'message': f'Successfully processed {result["total_records"]} records. Saved: {saved_count}, Updated: {updated_count}',
                'data': result['data']
            })
        else:
            # Clean up uploaded file
            try:
                os.remove(file_path)
            except:
                pass
            
            return jsonify(result), 400
            
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error uploading file: {str(e)}'}), 500

@plan_scraper_bp.route('/api/plan-scraper', methods=['GET'])
@login_required
def get_plan_scraper_data():
    """Get plan scraper data with pagination and filtering"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_get_plan_scraper_data_impl)()

def _get_plan_scraper_data_impl():
    try:
        db = get_db()
        PlanScraperData = get_plan_scraper_model()
        
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '', type=str)
        print_machine = request.args.get('print_machine', '', type=str)
        
        # DEBUG: Log parameters
        print(f"🔍 DEBUG: API called with parameters:")
        print(f"   page: {page}")
        print(f"   per_page: {per_page}")
        print(f"   search: '{search}'")
        print(f"   print_machine: '{print_machine}'")
        
        # Build query
        query = PlanScraperData.query
        
        # Apply filters
        if search:
            query = query.filter(
                or_(
                    PlanScraperData.wo_number.like(f'%{search}%'),
                    PlanScraperData.mc_number.like(f'%{search}%'),
                    PlanScraperData.item_name.like(f'%{search}%'),
                    PlanScraperData.paper_type.like(f'%{search}%')
                )
            )
            print(f"🔍 DEBUG: Applied search filter: '{search}'")
        
        if print_machine:
            query = query.filter(PlanScraperData.print_machine == print_machine)
            print(f"🔍 DEBUG: Applied machine filter: '{print_machine}'")
        
        # Count total records before pagination
        total_count = query.count()
        print(f"📊 DEBUG: Total records after filtering: {total_count}")
        
        # Order by created_at desc
        query = query.order_by(PlanScraperData.created_at.desc())
        
        # Paginate
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        items = pagination.items
        
        print(f"📄 DEBUG: Pagination results:")
        print(f"   Current page: {pagination.page}")
        print(f"   Total pages: {pagination.pages}")
        print(f"   Items per page: {per_page}")
        print(f"   Items returned: {len(items)}")
        print(f"   Has prev: {pagination.has_prev}")
        print(f"   Has next: {pagination.has_next}")
        
        # Debug: Show first few WO numbers
        if items:
            wo_numbers = [item.wo_number for item in items[:5]]
            print(f"   Sample WO numbers: {wo_numbers}")
        
        # Get WorkQueue model to check if plan scraper data has been received
        WorkQueue = get_work_queue_model()
        
        # Get all plan scraper data IDs that have work queue entries
        received_plan_ids = set()
        work_queue_entries = WorkQueue.query.all()
        for entry in work_queue_entries:
            if entry.plan_scraper_data_id:
                received_plan_ids.add(entry.plan_scraper_data_id)
        
        # Add received status to each item
        data_with_received = []
        for item in items:
            item_dict = item.to_dict()
            is_received = item.id in received_plan_ids
            item_dict['received'] = is_received
            data_with_received.append(item_dict)
        
        return jsonify({
            'success': True,
            'data': data_with_received,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_prev': pagination.has_prev,
                'has_next': pagination.has_next
            }
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _get_plan_scraper_data_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/plan-scraper/machines', methods=['GET'])
@login_required
def get_print_machines():
    """Get list of available print machines"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_get_print_machines_impl)()

def _get_print_machines_impl():
    try:
        db = get_db()
        PlanScraperData = get_plan_scraper_model()
        
        machines = db.session.query(PlanScraperData.print_machine).distinct().all()
        machine_list = [machine[0] for machine in machines if machine[0]]
        
        print(f"🖨️ DEBUG: Available machines: {machine_list}")
        
        return jsonify({
            'success': True,
            'data': machine_list
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _get_print_machines_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/plan-scraper/<int:id>', methods=['GET'])
@login_required
def get_plan_scraper_record(id):
    """Get a single plan scraper record"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_get_plan_scraper_record_impl)(id)

def _get_plan_scraper_record_impl(id):
    try:
        db = get_db()
        PlanScraperData = get_plan_scraper_model()
        WorkQueue = get_work_queue_model()
        
        record = PlanScraperData.query.get(id)
        if not record:
            return jsonify({'success': False, 'error': 'Record not found'}), 404
        
        # Check if this record has been received (has work queue entry)
        work_queue_entry = WorkQueue.query.filter_by(plan_scraper_data_id=id).first()
        received = work_queue_entry is not None
        
        # Add received status to the record
        record_dict = record.to_dict()
        record_dict['received'] = received
        
        return jsonify({
            'success': True,
            'data': record_dict
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _get_plan_scraper_record_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/plan-scraper/<int:id>', methods=['DELETE'])
@login_required
def delete_plan_scraper_record(id):
    """Delete a plan scraper record"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_delete_plan_scraper_record_impl)(id)

def _delete_plan_scraper_record_impl(id):
    try:
        print(f"🗑️ DEBUG: Attempting to delete record with ID: {id}")
        db = get_db()
        PlanScraperData = get_plan_scraper_model()
        
        record = PlanScraperData.query.get(id)
        if not record:
            print(f"❌ DEBUG: Record with ID {id} not found")
            return jsonify({'success': False, 'error': 'Record not found'}), 404
        
        print(f"📋 DEBUG: Found record to delete: {record.wo_number} - {record.print_machine}")
        
        db.session.delete(record)
        db.session.commit()
        
        print(f"✅ DEBUG: Successfully deleted record with ID: {id}")
        
        return jsonify({
            'success': True,
            'message': 'Record deleted successfully'
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _delete_plan_scraper_record_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/work-queue/create')
@login_required
def create_work_order_page():
    """Render halaman Create Work Order dari Work Queue"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_create_work_order_page_impl)()

def _create_work_order_page_impl():
    try:
        # Get work queue IDs from query parameters (support both single and multiple)
        work_queue_ids_param = request.args.get('work_queue_ids')
        work_queue_id = request.args.get('work_queue_id', type=int)
        complete_job = request.args.get('complete_job', 'false').lower() == 'true'
        
        # Handle both single ID and multiple IDs
        if work_queue_ids_param:
            # Multiple work queue IDs (comma-separated)
            try:
                work_queue_ids = [int(id.strip()) for id in work_queue_ids_param.split(',') if id.strip()]
            except (ValueError, AttributeError):
                flash('Invalid work queue IDs format', 'danger')
                return redirect(url_for('plan_scraper.work_queue_page'))
        elif work_queue_id:
            # Single work queue ID
            work_queue_ids = [work_queue_id]
        else:
            flash('Work Queue ID is required', 'danger')
            return redirect(url_for('plan_scraper.work_queue_page'))
        
        if not work_queue_ids:
            flash('Work Queue ID is required', 'danger')
            return redirect(url_for('plan_scraper.work_queue_page'))
        
        # Get work queue data
        db = get_db()
        WorkQueue = get_work_queue_model()
        
        # Handle multiple work queue IDs
        work_queue_items = []
        for wq_id in work_queue_ids:
            wq_item = WorkQueue.query.get(wq_id)
            if wq_item:
                work_queue_items.append(wq_item)
        
        if not work_queue_items:
            flash('Work Queue items not found', 'danger')
            return redirect(url_for('plan_scraper.work_queue_page'))
        
        # Use first work queue item as primary for form display
        work_queue_item = work_queue_items[0]
        
        # If single work queue ID is provided, also check for merged items
        if len(work_queue_ids) == 1 and work_queue_id:
            # Check if there are any items merged into this primary item
            merged_items = WorkQueue.query.filter_by(merged_with_id=work_queue_id).all()
            if merged_items:
                print(f"🔗 MERGE: Found {len(merged_items)} items merged into work queue {work_queue_id}")
                work_queue_items.extend(merged_items)
        
        # Check if plan_data exists for all work queue items
        for item in work_queue_items:
            if not item.plan_data:
                flash('Plan data not found for one of the work queue items', 'danger')
                return redirect(url_for('plan_scraper.work_queue_page'))
        
        # Check work order status based on mode
        # For complete job mode: allow 'active' or 'in_progress' status
        # For create job mode: only allow 'active' status
        for item in work_queue_items:
            if complete_job:
                # Complete job mode: allow active or in_progress
                if item.status not in ['active', 'in_progress']:
                    flash('All work orders must be active or in-progress to complete jobs', 'danger')
                    return redirect(url_for('plan_scraper.work_queue_page'))
            else:
                # Create job mode: only allow active
                if item.status != 'active':
                    flash('All work orders must be active to create jobs', 'danger')
                    return redirect(url_for('plan_scraper.work_queue_page'))
        
        # Check for duplicate MC numbers
        # Get MC number from current work queue item
        current_mc_number = work_queue_item.plan_data.mc_number
        
        # Find all PlanScraperData with the same MC number
        same_mc_plan_data_ids = db.session.query(get_plan_scraper_model().id).filter(
            get_plan_scraper_model().mc_number == current_mc_number
        ).all()
        
        # Extract IDs from the result
        same_mc_ids = [item[0] for item in same_mc_plan_data_ids]
        
        # Find WorkQueue items with those plan_scraper_data_ids (excluding current items)
        duplicate_mc_items = WorkQueue.query.filter(
            WorkQueue.plan_scraper_data_id.in_(same_mc_ids),
            WorkQueue.id.notin_([item.id for item in work_queue_items]),
            WorkQueue.status == 'active'
        ).all()
        
        # Get dropdown data using SQLAlchemy models
        try:
            production_models = get_production_models()
            print(f"🔍 DEBUG: Production models loaded: {list(production_models.keys())}")
            
            ProductionCustomerName = production_models['ProductionCustomerName']
            ProductionImpositionCockpit = production_models['ProductionImpositionCockpit']
            # CalibrationReference is from models.py, not production_models
            from models import CalibrationReference
            ProductionImpositionRemarks = production_models['ProductionImpositionRemarks']
            ProductionPrintMachine = production_models['ProductionPrintMachine']
            print(f"🔍 DEBUG: All production model classes loaded successfully")
        except Exception as e:
            print(f"🔍 DEBUG: Error loading production models: {e}")
            import traceback
            traceback.print_exc()
            production_models = {}
            ProductionCustomerName = None
            ProductionImpositionCockpit = None
            ProductionImpositionCalibration = None
            ProductionImpositionRemarks = None
        
        try:
            if ProductionCustomerName:
                customers = ProductionCustomerName.query.order_by(ProductionCustomerName.customer_name).all()
                print(f"✅ Successfully loaded {len(customers)} customers")
                for customer in customers[:3]:  # Show first 3 for debugging
                    print(f"   - Customer ID: {customer.id}, Name: {customer.customer_name}")
            else:
                print("❌ ProductionCustomerName model is None")
                customers = []
        except Exception as e:
            print(f"❌ Error getting customers: {e}")
            import traceback
            traceback.print_exc()
            customers = []
            
        try:
            if ProductionImpositionCockpit:
                cockpits = ProductionImpositionCockpit.query.order_by(ProductionImpositionCockpit.imposition_cockpit).all()
                print(f"✅ Successfully loaded {len(cockpits)} cockpits")
                for cockpit in cockpits[:3]:  # Show first 3 for debugging
                    print(f"   - Cockpit ID: {cockpit.id}, Name: {cockpit.imposition_cockpit}")
            else:
                print("❌ ProductionImpositionCockpit model is None")
                cockpits = []
        except Exception as e:
            print(f"❌ Error getting cockpits: {e}")
            import traceback
            traceback.print_exc()
            cockpits = []
        
        try:
            if ProductionPrintMachine:
                print_machines = ProductionPrintMachine.query.order_by(ProductionPrintMachine.print_machine).all()
                print(f"✅ Successfully loaded {len(print_machines)} print machines")
                for machine in print_machines[:3]:  # Show first 3 for debugging
                    print(f"   - Print Machine ID: {machine.id}, Name: {machine.print_machine}")
            else:
                print("❌ ProductionPrintMachine model is None")
                print_machines = []
        except Exception as e:
            print(f"❌ Error getting print machines: {e}")
            import traceback
            traceback.print_exc()
            print_machines = []
        
        try:
            # Get print machine from work queue item to filter calibrations
            print_machine_filter = None
            if work_queue_item and work_queue_item.plan_data and work_queue_item.plan_data.print_machine:
                print_machine_filter = work_queue_item.plan_data.print_machine
            
            print(f"🔍 DEBUG: Print machine filter for calibrations: {print_machine_filter}")
            print(f"🔍 DEBUG: Work queue item exists: {work_queue_item is not None}")
            print(f"🔍 DEBUG: Plan data exists: {work_queue_item.plan_data is not None if work_queue_item else False}")
            if work_queue_item and work_queue_item.plan_data:
                print(f"🔍 DEBUG: Print machine in plan data: {work_queue_item.plan_data.print_machine}")
            
            if print_machine_filter:
                # Filter calibrations by print machine
                print(f"🔍 DEBUG: Attempting to filter calibrations by print_machine: '{print_machine_filter}' (type: {type(print_machine_filter)})")
                
                # First, let's see what print machines exist in calibrations
                all_machines = db.session.query(CalibrationReference.print_machine).distinct().all()
                print(f"🔍 DEBUG: Available print machines in calibrations: {[m[0] for m in all_machines]}")
                
                calibrations = CalibrationReference.query.filter_by(print_machine=print_machine_filter).order_by(CalibrationReference.calib_name).all()
                print(f"🔍 DEBUG: Successfully loaded {len(calibrations)} calibrations for print machine: {print_machine_filter}")
            else:
                # Load all calibrations if no print machine filter
                calibrations = CalibrationReference.query.order_by(CalibrationReference.calib_name).all()
                print(f"🔍 DEBUG: Successfully loaded {len(calibrations)} calibrations (all)")
            
            for calibration in calibrations[:3]:  # Show first 3 for debugging
                print(f"   - Calibration ID: {calibration.id}, Name: {calibration.calib_name}, Machine: {calibration.print_machine}")
        except Exception as e:
            print(f"❌ Error getting calibrations: {e}")
            import traceback
            traceback.print_exc()
            calibrations = []
            
        try:
            if ProductionImpositionRemarks:
                remarks = ProductionImpositionRemarks.query.order_by(ProductionImpositionRemarks.imposition_remarks).all()
                print(f"✅ Successfully loaded {len(remarks)} remarks")
                for remark in remarks[:3]:  # Show first 3 for debugging
                    print(f"   - Remark ID: {remark.id}, Name: {remark.imposition_remarks}")
            else:
                print("❌ ProductionImpositionRemarks model is None")
                remarks = []
        except Exception as e:
            print(f"❌ Error getting remarks: {e}")
            import traceback
            traceback.print_exc()
            remarks = []
        
        # Get users for PIC dropdown
        try:
            def get_user_model():
                from app import User
                return User
            
            User = get_user_model()
            users = User.query.filter_by(is_active=True).order_by(User.name).all()
            print(f"✅ Successfully loaded {len(users)} users")
        except Exception as e:
            print(f"❌ Error getting users: {e}")
            users = []
        
        return render_template('plan_scraper/create_work_order.html',
                           work_queue_item=work_queue_item,
                           work_queue_items=work_queue_items,
                           work_queue_ids=work_queue_ids,
                           complete_job=complete_job,
                           duplicate_mc_items=duplicate_mc_items,
                           customers=customers,
                           cockpits=cockpits,
                           calibrations=calibrations,
                           remarks=remarks,
                           print_machines=print_machines,
                           users=users)
        
    except Exception as e:
        print(f"🚨 ERROR in _create_work_order_page_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading work order data: {str(e)}', 'danger')
        return redirect(url_for('plan_scraper.work_queue_page'))

@plan_scraper_bp.route('/api/work-queue/dropdown-data', methods=['GET'])
@login_required
def get_dropdown_data():
    """Get dropdown data for Create Work Order form"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_get_dropdown_data_impl)()

def _get_dropdown_data_impl():
    try:
        db = get_db()
        
        # Get dropdown data using SQLAlchemy models
        production_models = get_production_models()
        ProductionCustomerName = production_models['ProductionCustomerName']
        ProductionImpositionCockpit = production_models['ProductionImpositionCockpit']
        ProductionImpositionRemarks = production_models['ProductionImpositionRemarks']
        ProductionPrintMachine = production_models['ProductionPrintMachine']
        
        # CalibrationReference is from models.py, not production_models
        from models import CalibrationReference
        
        # Get users for PIC dropdown
        def get_user_model():
            from app import User
            return User
        
        User = get_user_model()
        
        try:
            customers = ProductionCustomerName.query.order_by(ProductionCustomerName.customer_name).all()
            print(f"✅ API: Successfully loaded {len(customers)} customers")
        except Exception as e:
            print(f"❌ API Error getting customers: {e}")
            customers = []
            
        try:
            cockpits = ProductionImpositionCockpit.query.order_by(ProductionImpositionCockpit.imposition_cockpit).all()
            print(f"✅ API: Successfully loaded {len(cockpits)} cockpits")
        except Exception as e:
            print(f"❌ API Error getting cockpits: {e}")
            cockpits = []
            
        try:
            calibrations = CalibrationReference.query.order_by(CalibrationReference.calib_name).all()
            print(f"✅ API: Successfully loaded {len(calibrations)} calibrations")
        except Exception as e:
            print(f"❌ API Error getting calibrations: {e}")
            calibrations = []
            
        try:
            remarks = ProductionImpositionRemarks.query.order_by(ProductionImpositionRemarks.imposition_remarks).all()
            print(f"✅ API: Successfully loaded {len(remarks)} remarks")
        except Exception as e:
            print(f"❌ API Error getting remarks: {e}")
            remarks = []
        
        try:
            if ProductionPrintMachine:
                print_machines = ProductionPrintMachine.query.order_by(ProductionPrintMachine.print_machine).all()
                print(f"✅ API: Successfully loaded {len(print_machines)} print machines")
            else:
                print("❌ ProductionPrintMachine model is None")
                print_machines = []
        except Exception as e:
            print(f"❌ API Error getting print machines: {e}")
            print_machines = []
        
        try:
            users = User.query.filter_by(is_active=True).order_by(User.name).all()
            print(f"✅ API: Successfully loaded {len(users)} users")
        except Exception as e:
            print(f"❌ API Error getting users: {e}")
            users = []
        
        return jsonify({
            'success': True,
            'data': {
                'customers': [customer.to_dict() for customer in customers],
                'cockpits': [cockpit.to_dict() for cockpit in cockpits],
                'calibrations': [calibration.to_dict() for calibration in calibrations],
                'remarks': [remark.to_dict() for remark in remarks],
                'print_machines': [machine.to_dict() for machine in print_machines],
                'users': [{'id': user.id, 'name': user.name} for user in users]
            }
        })
        
    except Exception as e:
        print(f"🚨 ERROR in _get_dropdown_data_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/calibrations', methods=['GET'])
@login_required
def get_calibrations_by_machine():
    """Get calibrations filtered by print machine"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_get_calibrations_by_machine_impl)()

def _get_calibrations_by_machine_impl():
    try:
        # Get print machine from query parameter
        print_machine = request.args.get('print_machine')
        print(f"🔍 DEBUG: API called with print_machine parameter: {print_machine}")
        
        if not print_machine:
            print("🔍 DEBUG: No print machine parameter provided")
            return jsonify({
                'success': False,
                'error': 'Print machine parameter is required'
            }), 400
        
        # CalibrationReference is from models.py
        from models import CalibrationReference
        
        # Filter calibrations by print machine
        calibrations = CalibrationReference.query.filter_by(print_machine=print_machine).order_by(CalibrationReference.calib_name).all()
        
        print(f"🔍 DEBUG: Found {len(calibrations)} calibrations for print machine: {print_machine}")
        for cal in calibrations[:3]:  # Show first 3 for debugging
            print(f"   - Calibration: {cal.calib_name} (ID: {cal.id})")
        
        return jsonify({
            'success': True,
            'data': [calibration.to_dict() for calibration in calibrations]
        })
        
    except Exception as e:
        print(f"🚨 ERROR in _get_calibrations_by_machine_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@plan_scraper_bp.route('/api/work-queue/check-duplicate-mc/<int:id>', methods=['GET'])
@login_required
def check_duplicate_mc_by_id(id):
    """Check for duplicate MC numbers by work queue ID"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_check_duplicate_mc_by_id_impl)(id)

def _check_duplicate_mc_by_id_impl(id):
    try:
        db = get_db()
        WorkQueue = get_work_queue_model()
        
        # Get the work queue item
        work_queue_item = WorkQueue.query.get(id)
        if not work_queue_item:
            return jsonify({'success': False, 'error': 'Work Queue item not found'}), 404
        
        # Get MC number from current work queue item
        if not work_queue_item.plan_data:
            return jsonify({'success': False, 'error': 'Plan data not found for this work queue item'}), 404
        
        current_mc_number = work_queue_item.plan_data.mc_number
        
        # Find all WorkQueue items with same MC number (excluding current one)
        duplicate_items = WorkQueue.query.filter(
            WorkQueue.plan_scraper_data_id.in_(
                db.session.query(get_plan_scraper_model().id).filter(
                    get_plan_scraper_model().mc_number == current_mc_number
                )
            ),
            WorkQueue.id != id,
            WorkQueue.status == 'active'
        ).all()
        
        return jsonify({
            'success': True,
            'data': {
                'has_duplicates': len(duplicate_items) > 0,
                'mc_number': current_mc_number,
                'original': work_queue_item.to_dict_with_plan_data(),
                'duplicates': [item.to_dict_with_plan_data() for item in duplicate_items]
            }
        })
        
    except Exception as e:
        print(f"🚨 ERROR in _check_duplicate_mc_by_id_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/check-duplicate-mc', methods=['POST'])
@login_required
def check_duplicate_mc():
    """Check for duplicate MC numbers"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_check_duplicate_mc_impl)()

def _check_duplicate_mc_impl():
    try:
        data = request.get_json()
        mc_number = data.get('mc_number')
        exclude_id = data.get('exclude_id')
        
        if not mc_number:
            return jsonify({'success': False, 'error': 'MC number is required'}), 400
        
        db = get_db()
        WorkQueue = get_work_queue_model()
        
        query = WorkQueue.query.filter(
            WorkQueue.mc_number == mc_number,
            WorkQueue.status == 'active'
        )
        
        if exclude_id:
            query = query.filter(WorkQueue.id != exclude_id)
        
        duplicate_items = query.all()
        
        return jsonify({
            'success': True,
            'data': {
                'has_duplicates': len(duplicate_items) > 0,
                'duplicates': [item.to_dict() for item in duplicate_items]
            }
        })
        
    except Exception as e:
        print(f"🚨 ERROR in _check_duplicate_mc_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/check-duplicate-active-mc/<int:id>', methods=['GET'])
@login_required
def check_duplicate_active_mc_by_id(id):
    """Check for duplicate active MC numbers by work queue ID (for start job flow)"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_check_duplicate_active_mc_by_id_impl)(id)

def _check_duplicate_active_mc_by_id_impl(id):
    """Check for duplicate active work orders with same MC number"""
    try:
        db = get_db()
        WorkQueue = get_work_queue_model()
        
        # Get the work queue item
        work_queue_item = WorkQueue.query.get(id)
        if not work_queue_item:
            return jsonify({'success': False, 'error': 'Work Queue item not found'}), 404
        
        # Get MC number from current work queue item
        if not work_queue_item.plan_data:
            return jsonify({'success': False, 'error': 'Plan data not found for this work queue item'}), 404
        
        current_mc_number = work_queue_item.plan_data.mc_number
        
        # Find all active WorkQueue items with same MC number (excluding current one)
        duplicate_items = WorkQueue.query.filter(
            WorkQueue.plan_scraper_data_id.in_(
                db.session.query(get_plan_scraper_model().id).filter(
                    get_plan_scraper_model().mc_number == current_mc_number
                )
            ),
            WorkQueue.id != id,
            WorkQueue.status == 'active'  # Only active work orders
        ).all()
        
        return jsonify({
            'success': True,
            'data': {
                'has_duplicates': len(duplicate_items) > 0,
                'mc_number': current_mc_number,
                'original': work_queue_item.to_dict_with_plan_data(),
                'duplicates': [item.to_dict_with_plan_data() for item in duplicate_items]
            }
        })
        
    except Exception as e:
        print(f"🚨 ERROR in _check_duplicate_active_mc_by_id_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/create-production-job', methods=['POST'])
@login_required
def create_production_job():
    """Create production job from work queue"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_create_production_job_impl)()

def _create_production_job_impl():
    try:
        data = request.get_json()
        work_queue_id = data.get('work_queue_id')
        work_queue_ids = data.get('work_queue_ids')
        complete_job = data.get('complete_job_mode', 'false').lower() == 'true'
        
        # Handle both single and multiple work queue IDs
        if work_queue_ids:
            # Multiple work queue IDs
            if not isinstance(work_queue_ids, list):
                return jsonify({'success': False, 'error': 'Work queue IDs must be an array'}), 400
        elif work_queue_id:
            # Single work queue ID (backward compatibility)
            work_queue_ids = [work_queue_id]
        else:
            return jsonify({'success': False, 'error': 'Work Queue ID is required'}), 400
        
        if not work_queue_ids:
            return jsonify({'success': False, 'error': 'Work Queue ID is required'}), 400
        
        db = get_db()
        WorkQueue = get_work_queue_model()
        ProductionImpositionJob = get_production_imposition_job_model()
        
        # Get all work queue items
        work_queue_items = WorkQueue.query.filter(WorkQueue.id.in_(work_queue_ids)).all()
        
        if not work_queue_items:
            return jsonify({'success': False, 'error': 'Work Queue items not found'}), 404
        
        # Check if all work orders are active/in_progress and have plan data
        for item in work_queue_items:
            if not item.plan_data:
                return jsonify({'success': False, 'error': 'Plan data not found for work queue item'}), 404
            if item.status not in ['active', 'in_progress']:
                return jsonify({'success': False, 'error': 'Only active or in-progress work orders can be created'}), 400
        
        # Use the first work queue item as the primary one for job creation
        work_queue_item = work_queue_items[0]
        
        # Find user by name and get their ID
        from models import User
        pic_name = data.get('pic')
        user = User.query.filter_by(name=pic_name).first()
        pic_id = user.id if user else None
        
        # Get customer name from customer_id
        from .production_models import ProductionCustomerName
        customer_id = int(data.get('customer_name')) if data.get('customer_name') else None
        customer = ProductionCustomerName.query.get(customer_id)
        customer_name = customer.customer_name if customer else None
        
        # Get cockpit name from cockpit_id
        from .production_models import ProductionImpositionCockpit
        cockpit_id = int(data.get('cockpit_id')) if data.get('cockpit_id') else None
        cockpit = ProductionImpositionCockpit.query.get(cockpit_id)
        cockpit_name = cockpit.imposition_cockpit if cockpit else None
        
        # Get calibration name from calibration_id
        from models import CalibrationReference
        calibration_id = int(data.get('calibration_id')) if data.get('calibration_id') else None
        calibration = CalibrationReference.query.get(calibration_id)
        calibration_name = calibration.calib_name if calibration else None
        
        # Get remarks name from remarks_id
        from .production_models import ProductionImpositionRemarks
        remarks_id = int(data.get('remarks_id')) if data.get('remarks_id') else None
        remarks = ProductionImpositionRemarks.query.get(remarks_id)
        remarks_name = remarks.imposition_remarks if remarks else None
        
        # Create production job(s)
        # For multiple work queue IDs with same MC, create ONE job with combined WO numbers
        production_models = get_production_models()
        ProductionImpositionJob = production_models['ProductionImpositionJob']
        
        # Combine WO numbers from all work queue items
        wo_numbers = [wq.plan_data.wo_number for wq in work_queue_items if wq.plan_data]
        combined_wo_number = ', '.join(wo_numbers)
        
        # Create single production job
        # Use started_at and started_by from work_queue_item (set during Start Job)
        # If not yet started (in complete_job mode without prior start), use current datetime
        production_job_started_at = work_queue_item.started_at if work_queue_item.started_at else datetime.now(jakarta_tz)
        production_job_started_by = work_queue_item.started_by if work_queue_item.started_by else int(current_user.id) if current_user.id else None
        
        production_job = ProductionImpositionJob(
            work_queue_id=int(work_queue_items[0].id) if work_queue_items[0].id else None,
            plan_scraper_data_id=int(work_queue_items[0].plan_scraper_data_id) if work_queue_items[0].plan_scraper_data_id else None,
            tanggal=datetime.strptime(data.get('tanggal'), '%Y-%m-%d').date(),
            pic=pic_id,
            grup=data.get('grup'),
            shift=data.get('shift'),
            customer_name=customer_name,
            customer_id=customer_id,
            wo_number=combined_wo_number,  # Combined WO numbers
            mc_number=work_queue_items[0].plan_data.mc_number,
            item_name=work_queue_items[0].plan_data.item_name,
            up=int(work_queue_items[0].plan_data.num_up) if work_queue_items[0].plan_data.num_up else 0,
            paper_desc=work_queue_items[0].plan_data.paper_desc,
            paper_type=work_queue_items[0].plan_data.paper_type,
            paper_size=data.get('paper_size'),
            file_name=data.get('file_name'),
            print_block=data.get('print_block'),
            print_machine=work_queue_items[0].plan_data.print_machine,
            cockpit=cockpit_name,
            cockpit_id=cockpit_id,
            tiff_b_usage=data.get('tiff_b_usage'),
            calibration_name=calibration_name,
            calibration_id=calibration_id,
            remarks=remarks_name,
            remarks_id=remarks_id,
            started_at=production_job_started_at,
            started_by=production_job_started_by
        )
        
        # Set completion info if complete_job mode
        if complete_job:
            production_job.completed_at = datetime.now(jakarta_tz)
            production_job.completed_by = int(current_user.id) if current_user.id else None
        
        db.session.add(production_job)
        
        # Update ALL work queue records to same status
        for wq_item in work_queue_items:
            if complete_job:
                wq_item.status = 'completed'
                print(f"✅ DEBUG: Setting work queue {wq_item.id} to completed")
                # Set completion info from current user
                wq_item.completed_at = datetime.now(jakarta_tz)
                wq_item.completed_by = int(current_user.id) if current_user.id else None
            else:
                wq_item.status = 'in_progress'
                print(f"✅ DEBUG: Setting work queue {wq_item.id} to in_progress")
                # Only set started_at/by if not already set (first time starting job)
                if not wq_item.started_at:
                    wq_item.started_at = datetime.now(jakarta_tz)
                    wq_item.started_by = int(current_user.id) if current_user.id else None
            
            wq_item.updated_at = datetime.now(jakarta_tz)
        
        db.session.commit()
        
        # Prepare response message
        if len(work_queue_ids) > 1:
            message = f'Successfully {"completed" if complete_job else "created"} 1 job for {len(work_queue_ids)} work orders'
        else:
            message = 'Production job completed successfully' if complete_job else 'Production job created successfully'
        
        return jsonify({
            'success': True,
            'message': message,
            'data': {
                'job': production_job.to_dict(),
                'work_queue_count': len(work_queue_ids),
                'wo_numbers': wo_numbers
            }
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"🚨 ERROR in _create_production_job_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/plan-scraper/<int:id>/receive', methods=['POST'])
@login_required
def receive_work_order(id):
    """Receive a work order and add it to work queue"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_receive_work_order_impl)(id)

def _receive_work_order_impl(id):
    try:
        print(f"📥 DEBUG: Attempting to receive work order with ID: {id}")
        db = get_db()
        PlanScraperData = get_plan_scraper_model()
        
        # Get plan scraper data
        plan_data = PlanScraperData.query.get(id)
        if not plan_data:
            print(f"❌ DEBUG: Plan scraper data with ID {id} not found")
            return jsonify({'success': False, 'error': 'Work order not found'}), 404
        
        # Check if already received
        WorkQueue = get_work_queue_model()
        existing_queue = WorkQueue.get_by_plan_data_id(id)
        if existing_queue:
            print(f"⚠️ DEBUG: Work order {plan_data.wo_number} already received")
            return jsonify({'success': False, 'error': 'Work order already received'}), 400
        
        # Check for duplicate MC+item_name with active status - GET ALL DUPLICATES
        duplicate_queues = WorkQueue.query.filter(
            WorkQueue.plan_scraper_data_id.in_(
                db.session.query(PlanScraperData.id).filter(
                    PlanScraperData.mc_number == plan_data.mc_number,
                    PlanScraperData.item_name == plan_data.item_name
                )
            ),
            WorkQueue.status == 'active'
        ).all()
        
        # Get request data
        request_data = request.get_json() or {}
        priority = request_data.get('priority', 'normal')
        notes = request_data.get('notes', '')
        
        # Validate priority
        if priority not in ['low', 'normal', 'high']:
            priority = 'normal'
        
        # If duplicates found, return merge confirmation response with ALL duplicates
        if duplicate_queues:
            print(f"⚠️ MERGE LOGIC: Found {len(duplicate_queues)} duplicate MC+item(s)")
            
            # Build list of all existing WOs to be merged
            existing_wos = [
                {
                    'id': dq.id,
                    'wo_number': dq.plan_data.wo_number if dq.plan_data else 'Unknown',
                    'priority': dq.priority or 'normal'
                }
                for dq in duplicate_queues
            ]
            
            print(f"📋 Existing WOs: {[wo['wo_number'] for wo in existing_wos]}")
            print(f"📋 New WO: {plan_data.wo_number}")
            
            return jsonify({
                'success': True,
                'merge_required': True,
                'primary_work_queue_id': duplicate_queues[0].id,
                'existing_work_orders': existing_wos,
                'new_wo_number': plan_data.wo_number,
                'new_plan_data_id': id,
                'mc_number': plan_data.mc_number,
                'item_name': plan_data.item_name,
                'message': f'{len(duplicate_queues) + 1} work orders with same MC ({plan_data.mc_number}) and Item ({plan_data.item_name}). Merge or create separately?'
            })
        
        # Create work queue entry
        work_queue_entry = WorkQueue(
            plan_scraper_data_id=id,
            received_by=current_user.id,
            priority=priority,
            notes=notes
        )
        
        db.session.add(work_queue_entry)
        db.session.commit()
        
        print(f"✅ DEBUG: Successfully received work order {plan_data.wo_number}")
        
        return jsonify({
            'success': True,
            'merge_required': False,
            'message': f'Work order {plan_data.wo_number} received successfully',
            'data': work_queue_entry.to_dict()
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _receive_work_order_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/merge', methods=['POST'])
@login_required
def merge_work_orders():
    """Merge two work orders by combining WO numbers"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_merge_work_orders_impl)()

def _merge_work_orders_impl():
    try:
        print(f"🔗 MERGE: Processing work order merge")
        db = get_db()
        PlanScraperData = get_plan_scraper_model()
        WorkQueue = get_work_queue_model()
        
        request_data = request.get_json() or {}
        existing_work_queue_id = request_data.get('existing_work_queue_id')
        new_plan_data_id = request_data.get('new_plan_data_id')
        should_merge = request_data.get('should_merge', True)
        
        # Get existing work queue entry
        existing_queue = WorkQueue.query.get(existing_work_queue_id)
        if not existing_queue:
            return jsonify({'success': False, 'error': 'Existing work order not found'}), 404
        
        # Get new plan scraper data
        new_plan_data = PlanScraperData.query.get(new_plan_data_id)
        if not new_plan_data:
            return jsonify({'success': False, 'error': 'New work order not found'}), 404
        
        if should_merge:
            # MERGE: Update existing work queue with combined WO numbers
            existing_wo = existing_queue.plan_data.wo_number
            new_wo = new_plan_data.wo_number
            combined_wo = f"{existing_wo}, {new_wo}"
            
            # Update existing work queue entry
            existing_queue.combined_wo_numbers = True  # Mark as merged
            
            # Update all related plan_scraper_data to reference the existing work_queue
            # Actually, we need to update the existing queue's plan_data associations
            # For now, just add note about the merge
            print(f"🔗 MERGE: Merging {existing_wo} + {new_wo} = {combined_wo}")
            
            # Create new work queue entry that references the new plan data, then merge them
            # Strategy: Create new work queue entry, then mark it as merged with existing
            new_queue = WorkQueue(
                plan_scraper_data_id=new_plan_data_id,
                received_by=current_user.id,
                priority=existing_queue.priority,
                notes=existing_queue.notes,
                status=existing_queue.status,
                merged_with_id=existing_work_queue_id
            )
            
            db.session.add(new_queue)
            db.session.commit()
            
            print(f"✅ MERGE: Successfully merged work orders")
            return jsonify({
                'success': True,
                'merged': True,
                'message': f'Work orders merged successfully: {combined_wo}',
                'existing_work_queue_id': existing_work_queue_id,
                'new_work_queue_id': new_queue.id,
                'combined_wo_number': combined_wo
            })
        else:
            # SEPARATE: Create new independent work queue entry
            new_queue = WorkQueue(
                plan_scraper_data_id=new_plan_data_id,
                received_by=current_user.id,
                priority=request_data.get('priority', 'normal'),
                notes=request_data.get('notes', '')
            )
            
            db.session.add(new_queue)
            db.session.commit()
            
            print(f"✅ SEPARATE: Work order created separately")
            return jsonify({
                'success': True,
                'merged': False,
                'message': f'Work order {new_plan_data.wo_number} created separately',
                'new_work_queue_id': new_queue.id,
                'data': new_queue.to_dict()
            })
        
    except Exception as e:
        print(f"🚨 MERGE: Exception in _merge_work_orders_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/plan-scraper/<int:id>/receive', methods=['DELETE'])
@login_required
def cancel_receive_work_order(id):
    """Cancel receive work order"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_cancel_receive_work_order_impl)(id)

def _cancel_receive_work_order_impl(id):
    try:
        print(f"🚫 DEBUG: Attempting to cancel receive work order with ID: {id}")
        db = get_db()
        WorkQueue = get_work_queue_model()
        
        # Get work queue entry
        work_queue_entry = WorkQueue.get_by_plan_data_id(id)
        if not work_queue_entry:
            print(f"❌ DEBUG: Work queue entry for plan data ID {id} not found")
            return jsonify({'success': False, 'error': 'Work order not found in queue'}), 404
        
        # Check if can be cancelled (only active work orders can be cancelled)
        if work_queue_entry.status != 'active':
            print(f"⚠️ DEBUG: Cannot cancel work order with status: {work_queue_entry.status}")
            return jsonify({'success': False, 'error': 'Cannot cancel work order that is not active'}), 400
        
        # Delete work queue entry
        db.session.delete(work_queue_entry)
        db.session.commit()
        
        print(f"✅ DEBUG: Successfully cancelled receive for work order ID: {id}")
        
        return jsonify({
            'success': True,
            'message': 'Work order receive cancelled successfully'
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _cancel_receive_work_order_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/work-queue')
@login_required
def work_queue_page():
    """Render halaman Work Queue"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(lambda: render_template('plan_scraper/work_queue.html'))()

@plan_scraper_bp.route('/api/work-queue', methods=['GET'])
@login_required
def get_work_queue_data():
    """Get work queue data with pagination and filtering"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_get_work_queue_data_impl)()

def _get_work_queue_data_impl():
    try:
        db = get_db()
        WorkQueue = get_work_queue_model()
        
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 25, type=int)
        status = request.args.get('status', '', type=str)
        priority = request.args.get('priority', '', type=str)
        machine = request.args.get('machine', '', type=str)
        search = request.args.get('search', '', type=str)
        
        # Build filters
        filters = {}
        if status:
            filters['status'] = status
        if priority:
            filters['priority'] = priority
        if machine:
            filters['machine'] = machine
        if search:
            filters['search'] = search
        
        # Get paginated results
        pagination = WorkQueue.get_active_work_orders(page=page, per_page=per_page, filters=filters)
        items = pagination.items
        
        # Process items to handle merged work orders
        processed_items = []
        processed_ids = set()  # Track items already processed
        search_mode = 'search' in filters
        
        for item in items:
            if item.id in processed_ids:
                continue
            
            # Skip items that have been merged into another work queue
            if item.merged_with_id is not None:
                # Find the primary item
                primary_item = WorkQueue.query.get(item.merged_with_id)
                if primary_item:
                    if primary_item.id not in processed_ids:
                        # Process the primary item FIRST when in search mode
                        if search_mode:
                            # Get all merged items for this primary
                            merged_items = WorkQueue.query.filter_by(merged_with_id=primary_item.id).all()
                            
                            # Combine WO numbers
                            wo_numbers = [primary_item.plan_data.wo_number] if primary_item.plan_data else []
                            for merged_item in merged_items:
                                if merged_item.plan_data:
                                    wo_numbers.append(merged_item.plan_data.wo_number)
                                    processed_ids.add(merged_item.id)
                            
                            combined_wo = ', '.join(wo_numbers)
                            
                            # Create item dict for primary
                            item_dict = primary_item.to_dict_with_plan_data()
                            item_dict['wo_number'] = combined_wo
                            item_dict['is_merged'] = True
                            item_dict['merged_count'] = len(merged_items)
                            
                            processed_items.append(item_dict)
                            processed_ids.add(primary_item.id)
                        else:
                            # Not search mode, just skip secondary
                            processed_ids.add(item.id)
                            continue
                    else:
                        processed_ids.add(item.id)
                        continue
                else:
                    processed_ids.add(item.id)
                    continue
            else:
                # Item is primary, check if there are merged items
                merged_items = WorkQueue.query.filter_by(merged_with_id=item.id).all()
                
                if merged_items:
                    # Combine WO numbers
                    wo_numbers = [item.plan_data.wo_number] if item.plan_data else []
                    for merged_item in merged_items:
                        if merged_item.plan_data:
                            wo_numbers.append(merged_item.plan_data.wo_number)
                            processed_ids.add(merged_item.id)
                    
                    combined_wo = ', '.join(wo_numbers)
                    
                    # Create item dict and update WO number
                    item_dict = item.to_dict_with_plan_data()
                    item_dict['wo_number'] = combined_wo
                    item_dict['is_merged'] = True
                    item_dict['merged_count'] = len(merged_items)
                    
                    processed_items.append(item_dict)
                    processed_ids.add(item.id)
                else:
                    # No merged items, just add as normal
                    item_dict = item.to_dict_with_plan_data()
                    item_dict['is_merged'] = False
                    item_dict['merged_count'] = 0
                    processed_items.append(item_dict)
                    processed_ids.add(item.id)
        
        return jsonify({
            'success': True,
            'data': processed_items,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_prev': pagination.has_prev,
                'has_next': pagination.has_next
            }
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _get_work_queue_data_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/<int:id>', methods=['GET'])
@login_required
def get_work_queue_record(id):
    """Get a single work queue record"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_get_work_queue_record_impl)(id)

def _get_work_queue_record_impl(id):
    try:
        db = get_db()
        WorkQueue = get_work_queue_model()
        ProductionImpositionJob = get_production_imposition_job_model()
        
        record = WorkQueue.query.get(id)
        if not record:
            return jsonify({'success': False, 'error': 'Work queue record not found'}), 404
        
        # Get the response data
        response_data = record.to_dict_with_plan_data()
        
        # Initialize production-related fields with default values
        response_data['customer_name'] = None
        response_data['calibration_name'] = None
        response_data['remarks'] = None
        
        # Check for merged work orders
        merged_items = WorkQueue.query.filter_by(merged_with_id=id).all()
        if merged_items:
            print(f"🔗 MERGE: Found {len(merged_items)} merged items for work queue {id}")
            response_data['is_merged'] = True
            response_data['merged_count'] = len(merged_items)
            response_data['primary_wo_number'] = record.plan_data.wo_number if record.plan_data else None
            
            # Build merged work order details
            merged_details = []
            for merged_item in merged_items:
                merged_details.append({
                    'id': merged_item.id,
                    'wo_number': merged_item.plan_data.wo_number if merged_item.plan_data else None,
                    'mc_number': merged_item.plan_data.mc_number if merged_item.plan_data else None,
                    'item_name': merged_item.plan_data.item_name if merged_item.plan_data else None,
                    'print_machine': merged_item.plan_data.print_machine if merged_item.plan_data else None
                })
            
            response_data['merged_work_orders'] = merged_details
        else:
            response_data['is_merged'] = False
            response_data['merged_count'] = 0
        
        # Check if there's a related ProductionImpositionJob
        production_job = ProductionImpositionJob.query.filter_by(work_queue_id=id).first()
        if production_job:
            # Include production job data, especially started_at/by and completed_at/by
            response_data['production_job'] = {
                'id': production_job.id,
                'started_at': production_job.started_at.strftime('%Y-%m-%d %H:%M:%S') if production_job.started_at else None,
                'started_by': production_job.started_by,
                'started_by_name': production_job.started_user.name if production_job.started_user else None,
                'completed_at': production_job.completed_at.strftime('%Y-%m-%d %H:%M:%S') if production_job.completed_at else None,
                'completed_by': production_job.completed_by,
                'completed_by_name': production_job.completed_user.name if production_job.completed_user else None
            }
            
            # Override with production job data if available
            if production_job.started_at:
                response_data['started_at'] = production_job.started_at.strftime('%Y-%m-%d %H:%M:%S')
                response_data['started_by'] = production_job.started_by
                response_data['started_by_name'] = production_job.started_user.name if production_job.started_user else None
            
            if production_job.completed_at:
                response_data['completed_at'] = production_job.completed_at.strftime('%Y-%m-%d %H:%M:%S')
                response_data['completed_by'] = production_job.completed_by
                response_data['completed_by_name'] = production_job.completed_user.name if production_job.completed_user else None
            
            # Include paper_size and other details from production_imposition_jobs
            if production_job.paper_size:
                response_data['paper_size'] = production_job.paper_size
            
            # Include customer_name, calibration_name, and remarks from production_imposition_jobs
            if production_job.customer_name:
                response_data['customer_name'] = production_job.customer_name
            if production_job.calibration_name:
                response_data['calibration_name'] = production_job.calibration_name
            if production_job.remarks:
                response_data['remarks'] = production_job.remarks
        
        return jsonify({
            'success': True,
            'data': response_data
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _get_work_queue_record_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/<int:id>', methods=['PUT'])
@login_required
def update_work_queue_record(id):
    """Update a work queue record"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_update_work_queue_record_impl)(id)

def _update_work_queue_record_impl(id):
    try:
        print(f"📝 DEBUG: Updating work queue record with ID: {id}")
        db = get_db()
        WorkQueue = get_work_queue_model()
        
        record = WorkQueue.query.get(id)
        if not record:
            print(f"❌ DEBUG: Work queue record with ID {id} not found")
            return jsonify({'success': False, 'error': 'Work queue record not found'}), 404
        
        # Get request data
        request_data = request.get_json()
        if not request_data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        # Update fields
        if 'status' in request_data:
            new_status = request_data['status']
            if new_status in ['active', 'pending', 'completed', 'cancelled']:
                if new_status == 'completed' and record.status != 'completed':
                    record.mark_completed(current_user.id, request_data.get('notes'))
                elif new_status == 'cancelled':
                    record.cancel(request_data.get('notes'))
                elif new_status == 'pending':
                    # Handle pending status with reason
                    if 'downtime_reason' in request_data:
                        record.start_downtime(
                            reason=request_data['downtime_reason'],
                            notes=request_data.get('downtime_notes'),
                            user_id=current_user.id
                        )
                    else:
                        return jsonify({'success': False, 'error': 'Downtime reason is required when setting status to pending'}), 400
                elif new_status == 'active':
                    # End current downtime when returning to active
                    record.end_downtime(user_id=current_user.id)
                else:
                    record.status = new_status
        
        if 'priority' in request_data:
            priority = request_data['priority']
            if priority in ['low', 'normal', 'high']:
                record.priority = priority
        
        if 'notes' in request_data and request_data['notes'] is not None:
            record.notes = request_data['notes']
        
        db.session.commit()
        
        print(f"✅ DEBUG: Successfully updated work queue record ID: {id}")
        
        return jsonify({
            'success': True,
            'message': 'Work queue record updated successfully',
            'data': record.to_dict_with_plan_data()
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _update_work_queue_record_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/<int:id>/start-job', methods=['POST'])
@login_required
def start_job(id):
    """Start a job by updating status and recording start time"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_start_job_impl)(id)

def _start_job_impl(id):
    """Implementation for starting a job"""
    try:
        print(f"🚀 DEBUG: Starting job for work queue ID: {id}")
        db = get_db()
        WorkQueue = get_work_queue_model()
        ProductionImpositionJob = get_production_imposition_job_model()
        
        # Get work queue record
        work_queue = WorkQueue.query.get(id)
        if not work_queue:
            print(f"❌ DEBUG: Work queue record with ID {id} not found")
            return jsonify({'success': False, 'error': 'Work queue record not found'}), 404
        
        # Check if job is already started
        if work_queue.status == 'in_progress':
            print(f"⚠️ DEBUG: Job already started for work queue ID: {id}")
            return jsonify({'success': False, 'error': 'Job already started'}), 400
        
        # Find all merged work orders that are merged INTO this one
        merged_items = WorkQueue.query.filter_by(merged_with_id=id).all()
        print(f"🔗 MERGE: Found {len(merged_items)} items merged into work queue {id}")
        
        # Collect all work queue items to start  (primary + merged)
        all_items_to_start = [work_queue] + merged_items
        
        # Update all work queue items to in_progress
        for item in all_items_to_start:
            item.status = 'in_progress'
            item.started_at = datetime.now(jakarta_tz)
            item.started_by = current_user.id
            item.updated_at = datetime.now(jakarta_tz)
            print(f"📝 DEBUG: Updated work queue {item.id} to in_progress with started_at and started_by")
        
        db.session.commit()
        
        print(f"✅ DEBUG: Successfully started {len(all_items_to_start)} job(s) for work queue ID: {id}")
        
        return jsonify({
            'success': True,
            'message': f'Job started successfully ({len(all_items_to_start)} work order{"s" if len(all_items_to_start) > 1 else ""})'
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _start_job_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/start-jobs', methods=['POST'])
@login_required
def start_jobs_batch():
    """Start multiple jobs (batch) - creates one ProductionImpositionJob for multiple work orders"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_start_jobs_batch_impl)()

def _start_jobs_batch_impl():
    """Implementation for batch starting jobs"""
    try:
        data = request.get_json()
        work_queue_ids = data.get('work_queue_ids', [])
        started_by = data.get('started_by', current_user.name if current_user else 'Unknown')
        
        if not work_queue_ids or len(work_queue_ids) == 0:
            print(f"❌ DEBUG: No work queue IDs provided")
            return jsonify({'success': False, 'error': 'No work queue IDs provided'}), 400
        
        print(f"🚀 DEBUG: Starting batch job for work queue IDs: {work_queue_ids}")
        
        db = get_db()
        WorkQueue = get_work_queue_model()
        ProductionImpositionJob = get_production_imposition_job_model()
        
        # Fetch all work queue records
        work_queues = WorkQueue.query.filter(WorkQueue.id.in_(work_queue_ids)).all()
        
        if len(work_queues) != len(work_queue_ids):
            print(f"❌ DEBUG: Some work queue records not found")
            return jsonify({'success': False, 'error': 'Some work queue records not found'}), 404
        
        # Verify all are active status
        non_active = [wq for wq in work_queues if wq.status != 'active']
        if non_active:
            print(f"⚠️ DEBUG: Some work orders are not active: {[wq.id for wq in non_active]}")
            return jsonify({'success': False, 'error': f'Some work orders are not active'}), 400
        
        # Get MC number from first work queue
        if not work_queues[0].plan_data:
            return jsonify({'success': False, 'error': 'Plan data not found'}), 400
        
        mc_number = work_queues[0].plan_data.mc_number
        
        # Verify all work queues have the same MC number
        for wq in work_queues[1:]:
            if not wq.plan_data or wq.plan_data.mc_number != mc_number:
                print(f"❌ DEBUG: Work queues have different MC numbers")
                return jsonify({'success': False, 'error': 'All work orders must have the same MC number'}), 400
        
        # Update all work queue statuses to in_progress
        for wq in work_queues:
            wq.status = 'in_progress'
            wq.started_at = datetime.now(jakarta_tz)
            wq.started_by = current_user.id
            wq.updated_at = datetime.now(jakarta_tz)
            print(f"📝 DEBUG: Updated work queue {wq.id} to in_progress with started_at and started_by")
        
        db.session.commit()
        
        print(f"✅ DEBUG: Successfully started batch job for {len(work_queue_ids)} work orders with MC: {mc_number}")
        
        return jsonify({
            'success': True,
            'message': f'Successfully started {len(work_queue_ids)} work order(s)',
            'data': {
                'work_queue_ids': work_queue_ids,
                'mc_number': mc_number,
                'count': len(work_queue_ids)
            }
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _start_jobs_batch_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/<int:id>', methods=['DELETE'])
@login_required
def delete_work_queue_record(id):
    """Delete a work queue record"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_delete_work_queue_record_impl)(id)

def _delete_work_queue_record_impl(id):
    try:
        print(f"🗑️ DEBUG: Deleting work queue record with ID: {id}")
        db = get_db()
        WorkQueue = get_work_queue_model()
        
        record = WorkQueue.query.get(id)
        if not record:
            print(f"❌ DEBUG: Work queue record with ID {id} not found")
            return jsonify({'success': False, 'error': 'Work queue record not found'}), 404
        
        db.session.delete(record)
        db.session.commit()
        
        print(f"✅ DEBUG: Successfully deleted work queue record ID: {id}")
        
        return jsonify({
            'success': True,
            'message': 'Work queue record deleted successfully'
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _delete_work_queue_record_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/pending-reasons', methods=['GET'])
@login_required
def get_pending_reasons():
    """Get list of available pending reasons"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_get_pending_reasons_impl)()

def _get_pending_reasons_impl():
    try:
        # Only 5 reasons as requested by user
        pending_reasons = [
            {'value': 'SERVER_ERROR', 'label': 'Server Error'},
            {'value': 'PRINECT_ERROR', 'label': 'Prinect Error'},
            {'value': 'TUNGGU_DATA_PDND', 'label': 'Tunggu Data dari PDND'},
            {'value': 'TUNGGU_MILLAR_PDND', 'label': 'Tunggu Millar dari PDND'},
            {'value': 'TUNGGU_DIGITAL_PRINT_PDND', 'label': 'Tunggu Digital Print dari PDND'}
        ]
        
        return jsonify({
            'success': True,
            'data': pending_reasons
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _get_pending_reasons_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/machines', methods=['GET'])
@login_required
def get_work_queue_machines():
    """Get list of available print machines for work queue"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_get_work_queue_machines_impl)()

def _get_work_queue_machines_impl():
    try:
        db = get_db()
        WorkQueue = get_work_queue_model()
        PlanScraperData = get_plan_scraper_model()
        
        # Get distinct machines from work queue entries
        machines = db.session.query(PlanScraperData.print_machine)\
            .join(WorkQueue, PlanScraperData.id == WorkQueue.plan_scraper_data_id)\
            .distinct()\
            .all()
        
        machine_list = [machine[0] for machine in machines if machine[0]]
        
        print(f"🖨️ DEBUG: Available work queue machines: {machine_list}")
        
        return jsonify({
            'success': True,
            'data': machine_list
        })
        
    except Exception as e:
        print(f"🚨 DEBUG: Exception in _get_work_queue_machines_impl: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/<int:id>/downtime', methods=['POST'])
@login_required
def start_downtime(id):
    """Start downtime for a work order"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_start_downtime_impl)(id)

def _start_downtime_impl(id):
    try:
        db = get_db()
        WorkQueue = get_work_queue_model()
        
        record = WorkQueue.query.get(id)
        if not record:
            return jsonify({'success': False, 'error': 'Work queue record not found'}), 404
        
        # Handle merged work orders: if passed ID is secondary, get primary
        primary_record = record
        if record.merged_with_id:
            # This is a secondary (merged) WO, get the primary
            primary_record = WorkQueue.query.get(record.merged_with_id)
            if not primary_record:
                return jsonify({'success': False, 'error': 'Primary work queue record not found'}), 404
        
        # Get request data
        data = request.get_json()
        downtime_reason = data.get('downtime_reason')
        downtime_notes = data.get('downtime_notes', '')
        
        if not downtime_reason:
            return jsonify({'success': False, 'error': 'Downtime reason is required'}), 400
        
        # Start downtime on PRIMARY (which will affect all merged items)
        primary_record.start_downtime(downtime_reason, downtime_notes, current_user.id)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Downtime started successfully for all affected work orders',
            'data': primary_record.to_dict_with_plan_data()
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@plan_scraper_bp.route('/api/work-queue/<int:id>/downtime/end', methods=['POST'])
@login_required
def end_downtime(id):
    """End downtime for a work order"""
    require_mounting_access = get_require_mounting_access()
    return require_mounting_access(_end_downtime_impl)(id)

def _end_downtime_impl(id):
    try:
        db = get_db()
        WorkQueue = get_work_queue_model()
        
        record = WorkQueue.query.get(id)
        if not record:
            return jsonify({'success': False, 'error': 'Work queue record not found'}), 404
        
        # Handle merged work orders: if passed ID is secondary, get primary
        primary_record = record
        if record.merged_with_id:
            # This is a secondary (merged) WO, get the primary
            primary_record = WorkQueue.query.get(record.merged_with_id)
            if not primary_record:
                return jsonify({'success': False, 'error': 'Primary work queue record not found'}), 404
        
        # End downtime on PRIMARY (which will affect all merged items)
        primary_record.end_downtime(current_user.id)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Downtime ended successfully for all affected work orders',
            'data': primary_record.to_dict_with_plan_data()
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ===============================
# Print Prepress Form Route
# ===============================
@plan_scraper_bp.route('/api/print-prepress-form/<int:work_queue_id>', methods=['GET'])
@login_required
def print_prepress_form(work_queue_id):
    """Display printable prepress form for completed work order"""
    try:
        require_mounting_access = get_require_mounting_access()
        
        db = get_db()
        WorkQueue = get_work_queue_model()
        ProductionImpositionJob = get_production_imposition_job_model()
        
        # Fetch work queue record
        work_queue = WorkQueue.query.get(work_queue_id)
        if not work_queue:
            return jsonify({'success': False, 'error': 'Work queue record not found'}), 404
        
        # Fetch related production job
        production_job = ProductionImpositionJob.query.filter_by(work_queue_id=work_queue_id).first()
        
        # Prepare context data
        context = {
            'work_queue': work_queue,
            'production_job': production_job,
            'current_date': datetime.now(jakarta_tz).strftime('%d-%m-%Y'),
        }
        
        # Populate form fields from both models
        if production_job:
            context.update({
                'tanggal': production_job.tanggal.strftime('%d-%m-%Y') if production_job.tanggal else datetime.now(jakarta_tz).strftime('%d-%m-%Y'),
                'customer_name': production_job.customer_name,
                'item_name': production_job.item_name,
                'file_name': production_job.file_name,
                'print_block': production_job.print_block,
                'wo_number': production_job.wo_number,
                'paper_size': production_job.paper_size,
                'up': production_job.up,
                'print_machine': production_job.print_machine,
                'calibration_name': production_job.calibration_name,
                'completed_at': production_job.completed_at.strftime('%d-%m-%Y %H:%M') if production_job.completed_at else '',
                'completed_by_name': production_job.completed_user.name if production_job.completed_user else '',
            })
        else:
            # Fallback to work queue data if no production job
            context.update({
                'tanggal': work_queue.tanggal.strftime('%d-%m-%Y') if hasattr(work_queue, 'tanggal') and work_queue.tanggal else datetime.now(jakarta_tz).strftime('%d-%m-%Y'),
                'customer_name': work_queue.customer_name or '',
                'item_name': work_queue.item_name or '',
                'file_name': work_queue.file_name or '',
                'print_block': getattr(work_queue, 'print_block', '') or '',
                'wo_number': work_queue.wo_number or '',
                'paper_size': getattr(work_queue, 'paper_size', '') or '',
                'up': getattr(work_queue, 'up', '') or '',
                'print_machine': work_queue.print_machine or '',
                'calibration_name': getattr(work_queue, 'calibration_name', '') or '',
                'completed_at': work_queue.completed_at.strftime('%d-%m-%Y %H:%M') if work_queue.completed_at else '',
                'completed_by_name': work_queue.completed_user.name if hasattr(work_queue, 'completed_user') and work_queue.completed_user else '',
            })
        
        return render_template('plan_scraper/print_prepress_form.html', **context)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Error: {str(e)}", 500


@plan_scraper_bp.route('/api/print-raster-prepress-form/<int:work_queue_id>', methods=['GET'])
@login_required
def print_raster_prepress_form(work_queue_id):
    """Display printable raster prepress form for completed work order"""
    try:
        require_mounting_access = get_require_mounting_access()
        
        db = get_db()
        WorkQueue = get_work_queue_model()
        ProductionImpositionJob = get_production_imposition_job_model()
        
        # Fetch work queue record
        work_queue = WorkQueue.query.get(work_queue_id)
        if not work_queue:
            return jsonify({'success': False, 'error': 'Work queue record not found'}), 404
        
        # Fetch related production job
        production_job = ProductionImpositionJob.query.filter_by(work_queue_id=work_queue_id).first()
        
        # Prepare context data
        context = {
            'work_queue': work_queue,
            'production_job': production_job,
            'current_date': datetime.now(jakarta_tz).strftime('%d-%m-%Y'),
        }
        
        # Populate form fields from both models
        if production_job:
            context.update({
                'tanggal': production_job.tanggal.strftime('%d-%m-%Y') if production_job.tanggal else datetime.now(jakarta_tz).strftime('%d-%m-%Y'),
                'customer_name': production_job.customer_name,
                'item_name': production_job.item_name,
                'file_name': production_job.file_name,
                'print_block': production_job.print_block,
                'wo_number': production_job.wo_number,
                'paper_size': production_job.paper_size,
                'up': production_job.up,
                'print_machine': production_job.print_machine,
                'calibration_name': production_job.calibration_name,
                'completed_at': production_job.completed_at.strftime('%d-%m-%Y %H:%M') if production_job.completed_at else '',
                'completed_by_name': production_job.completed_user.name if production_job.completed_user else '',
            })
        else:
            # Fallback to work queue data if no production job
            context.update({
                'tanggal': work_queue.tanggal.strftime('%d-%m-%Y') if hasattr(work_queue, 'tanggal') and work_queue.tanggal else datetime.now(jakarta_tz).strftime('%d-%m-%Y'),
                'customer_name': work_queue.customer_name or '',
                'item_name': work_queue.item_name or '',
                'file_name': work_queue.file_name or '',
                'print_block': getattr(work_queue, 'print_block', '') or '',
                'wo_number': work_queue.wo_number or '',
                'paper_size': getattr(work_queue, 'paper_size', '') or '',
                'up': getattr(work_queue, 'up', '') or '',
                'print_machine': work_queue.print_machine or '',
                'calibration_name': getattr(work_queue, 'calibration_name', '') or '',
                'completed_at': work_queue.completed_at.strftime('%d-%m-%Y %H:%M') if work_queue.completed_at else '',
                'completed_by_name': work_queue.completed_user.name if hasattr(work_queue, 'completed_user') and work_queue.completed_user else '',
            })
        
        return render_template('plan_scraper/print_raster_prepress_form.html', **context)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Error: {str(e)}", 500


# Helper function to get WorkQueue model
def get_work_queue_model():
    from .models import WorkQueue
    return WorkQueue