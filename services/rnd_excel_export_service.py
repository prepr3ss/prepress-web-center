import os
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, or_
from models import db, User
from models_rnd import (
    RNDJob, RNDProgressStep, RNDJobProgressAssignment
)

# Jakarta timezone
jakarta_tz = None  # Will be set when service is initialized

class RNDExcelExportService:
    """Service for exporting RND job data to Excel format"""
    
    def __init__(self, db_session, timezone=None):
        self.db = db_session
        global jakarta_tz
        jakarta_tz = timezone or jakarta_tz
    
    def export_jobs_to_excel(self, filters=None):
        """
        Export RND jobs to Excel with comprehensive data including progress steps
        
        Args:
            filters (dict): Optional filters for date range, sample type, status
            
        Returns:
            io.BytesIO: Excel file buffer
        """
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "RND Jobs Export"
            
            # Get filtered jobs
            jobs_query = self._build_jobs_query(filters)
            jobs = jobs_query.all()
            
            if not jobs:
                # Raise exception instead of creating empty sheet
                raise ValueError("No jobs found matching the selected filters")
            
            # Get all progress assignments for these jobs
            job_ids = [job.id for job in jobs]
            progress_assignments = self.db.query(RNDJobProgressAssignment).filter(
                RNDJobProgressAssignment.job_id.in_(job_ids)
            ).all()
            
            # Group progress assignments by job
            job_progress_map = {}
            max_progress_steps = 0
            
            for assignment in progress_assignments:
                job_id = assignment.job_id
                if job_id not in job_progress_map:
                    job_progress_map[job_id] = []
                job_progress_map[job_id].append(assignment)
            
            # Find maximum number of progress steps for column generation
            for job_id, assignments in job_progress_map.items():
                # Sort assignments by step order
                sorted_assignments = sorted(assignments, key=lambda x: x.progress_step.step_order if x.progress_step else 999)
                job_progress_map[job_id] = sorted_assignments
                max_progress_steps = max(max_progress_steps, len(sorted_assignments))
            
            # Create headers
            self._create_headers(ws, max_progress_steps)
            
            # Add data rows
            row_num = 2  # Start from row 2 (after headers)
            
            for i, job in enumerate(jobs, start=1):
                row_data = self._create_job_row(job, job_progress_map.get(job.id, []), max_progress_steps, i)
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    
                    # Apply styling based on column type
                    self._apply_cell_styling(cell, col_num, value)
                
                row_num += 1
            
            # Apply column formatting
            self._apply_column_formatting(ws, max_progress_steps)
            
            # Add filters to all columns
            ws.auto_filter.ref = f"A1:{get_column_letter(10 + max_progress_steps * 4)}1"
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            return self._save_to_buffer(wb)
            
        except Exception as e:
            print(f"Error in export_jobs_to_excel: {str(e)}")
            raise e
    
    def _build_jobs_query(self, filters):
        """Build jobs query with optional filters"""
        query = RNDJob.query
        
        if filters:
            # Date range filter
            if filters.get('start_date'):
                start_date = filters['start_date']
                query = query.filter(RNDJob.started_at >= start_date)
            
            if filters.get('end_date'):
                end_date = filters['end_date']
                query = query.filter(RNDJob.started_at <= end_date)
            
            # Sample type filter
            if filters.get('sample_type'):
                query = query.filter(RNDJob.sample_type == filters['sample_type'])
            
            # Status filter
            if filters.get('status'):
                query = query.filter(RNDJob.status == filters['status'])
        
        # Order by started date descending
        query = query.order_by(RNDJob.started_at.desc())
        
        return query
    
    def _create_empty_sheet(self, ws):
        """Create empty sheet with headers"""
        headers = [
            "No.", "Started At", "RND Job ID", "Deadline", "Finished At",
            "Status", "Item Name", "Sample Type", "Priority Level", "Notes"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            self._apply_header_style(cell)
        
        # Add a message row indicating no data
        message_row = [
            "No data found", "", "", "", "", "", "", "", "", "", "", ""
        ]
        
        for col_num, value in enumerate(message_row, 1):
            cell = ws.cell(row=2, column=col_num, value=value)
            # Apply styling for message
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_num == 1:
                cell.font = Font(italic=True, color="666666")
    
    def _create_headers(self, ws, max_progress_steps):
        """Create comprehensive headers including dynamic progress step columns"""
        # Base job information headers
        base_headers = [
            "No.", "Started At", "RND Job ID", "Deadline", "Finished At", 
            "Status", "Item Name", "Sample Type", "Priority Level", "Notes"
        ]
        
        # Add dynamic progress step headers
        progress_headers = []
        for i in range(1, max_progress_steps + 1):
            progress_headers.extend([
                f"Progress Step {i}",
                f"PIC Progress Step {i}",
                f"Started At Progress Step {i}",
                f"Finished At Progress Step {i}"
            ])
        
        all_headers = base_headers + progress_headers
        
        # Apply headers to worksheet
        for col_num, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            self._apply_header_style(cell)
    
    def _create_job_row(self, job, progress_assignments, max_progress_steps, row_number):
        """Create a data row for a single job"""
        # Base job information
        row_data = [
            row_number,  # No. - filled with actual row number
            self._format_datetime(job.started_at),
            job.job_id,
            self._format_datetime(job.deadline_at),
            self._format_datetime(job.finished_at) or "Not Finished Yet",
            job.status,
            job.item_name,
            job.sample_type,
            job.priority_level,
            job.notes or ""
        ]
        
        # Progress step data
        progress_data = []
        for i in range(max_progress_steps):
            if i < len(progress_assignments):
                assignment = progress_assignments[i]
                progress_data.extend([
                    assignment.progress_step.name if assignment.progress_step else "Unknown",
                    assignment.pic.name if assignment.pic else "Unassigned",
                    self._format_datetime(assignment.started_at),
                    self._format_datetime(assignment.finished_at) or "Not Finished Yet"
                ])
            else:
                # Empty columns for jobs with fewer progress steps
                progress_data.extend(["", "", "", ""])
        
        return row_data + progress_data
    
    def _apply_header_style(self, cell):
        """Apply styling to header cells"""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def _apply_cell_styling(self, cell, col_num, value):
        """Apply styling to data cells based on column type and value"""
        # Center align certain columns
        if col_num in [1, 6]:  # No. and Status
            cell.alignment = Alignment(horizontal="center")
        
        # Apply color coding for status
        if col_num == 6 and isinstance(value, str):  # Status column
            if value == "completed":
                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            elif value == "in_progress":
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            elif value == "rejected":
                cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        # Apply border to all cells
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def _apply_column_formatting(self, ws, max_progress_steps):
        """Apply column width formatting"""
        # Base column widths
        column_widths = {
            'A': 8,   # No.
            'B': 20,  # Started At
            'C': 15,  # RND Job ID
            'D': 20,  # Deadline
            'E': 20,  # Finished At
            'F': 15,  # Status
            'G': 30,  # Item Name
            'H': 15,  # Sample Type
            'I': 15,  # Priority Level
            'J': 40   # Notes
        }
        
        # Apply base column widths
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Apply progress step column widths
        start_col = 11  # Column K
        for i in range(max_progress_steps):
            for j in range(4):  # 4 columns per progress step
                col_letter = get_column_letter(start_col + i * 4 + j)
                if j in [0, 1]:  # Step name and PIC name
                    ws.column_dimensions[col_letter].width = 25
                else:  # Date columns
                    ws.column_dimensions[col_letter].width = 20
    
    def _format_datetime(self, dt):
        """Format datetime for Excel display"""
        if dt is None:
            return None
        
        if hasattr(dt, 'strftime'):
            return dt.strftime('%d-%m-%Y %H:%M:%S')
        else:
            # Handle timezone-aware datetimes
            try:
                if dt.tzinfo is None:
                    # Assume Jakarta timezone for naive datetimes
                    import pytz
                    dt = jakarta_tz.localize(dt) if jakarta_tz else dt
                return dt.strftime('%d-%m-%Y %H:%M:%S')
            except:
                return str(dt)
    
    def _save_to_buffer(self, workbook):
        """Save workbook to in-memory buffer"""
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_filename(self):
        """Generate filename with timestamp"""
        timestamp = datetime.now().strftime('%d-%m-%Y_%H-%M-%S')
        return f"RND_Job_Export_{timestamp}.xlsx"