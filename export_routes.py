# Standard library imports
from datetime import datetime, time, timedelta
from functools import wraps
from io import BytesIO, StringIO
from urllib.parse import quote_plus
import calendar
import csv
import io
import locale
import os
import pytz
import random
import traceback

# Third party imports
from flask import Blueprint, abort, flash, jsonify, make_response, redirect, render_template, request, send_file, send_from_directory, session, url_for
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import String, and_, cast, extract, func, literal_column, or_, text
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
import pymysql

# Local imports
from config import DB_CONFIG
from models import db, Division, User, CTPProductionLog, PlateAdjustmentRequest, PlateBonRequest, KartuStockPlateFuji, KartuStockPlateSaphira, KartuStockChemicalFuji, KartuStockChemicalSaphira, MonthlyWorkHours, ChemicalBonCTP, BonPlate, CTPMachine, CTPProblemLog, CTPProblemPhoto, CTPProblemDocument
from plate_mappings import PlateTypeMapping

# Timezone untuk Jakarta
jakarta_tz = pytz.timezone('Asia/Jakarta')
now_jakarta = datetime.now(jakarta_tz)

# Helper function to format datetime in Indonesian
def format_datetime_indonesia(dt):
    """Format datetime to Indonesian format"""
    if not dt:
        return ''
    
    bulan_indonesia = [
        'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
        'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
    ]
    
    day = dt.day
    month = bulan_indonesia[dt.month - 1]
    year = dt.year
    hour = str(dt.hour).zfill(2)
    minute = str(dt.minute).zfill(2)
    
    return f"{day} {month} {year} - {hour}:{minute}"

def format_tanggal_indonesia(dt):
    """Format date to Indonesian format"""
    if not dt:
        return ''
    
    bulan_indonesia = [
        'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
        'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
    ]
    
    day = dt.day
    month = bulan_indonesia[dt.month - 1]
    year = dt.year
    
    return f"{day} {month} {year}"

# --- Export Routes ---

# Create Blueprint for export routes
export_bp = Blueprint('export', __name__)

# --- Export Routes ---
@export_bp.route('/export-ctp-logs')
@login_required
def export_ctp_logs():
    try:
        
        # Get filter parameters
        machine_nickname = request.args.get('machine_nickname')
        format_type = request.args.get('format', 'excel')  # Default to Excel
        year = request.args.get('year')
        month = request.args.get('month')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        technician_type = request.args.get('technician_type')
        status = request.args.get('status')
        search = request.args.get('search', '').strip()
        
        if not machine_nickname:
            return jsonify({'success': False, 'error': 'Machine nickname is required'}), 400
        
        # Get machine information
        machine = CTPMachine.query.filter_by(nickname=machine_nickname).first()
        if not machine:
            return jsonify({'success': False, 'error': 'Machine not found'}), 404
        
        # Build query with filters
        query = CTPProblemLog.query.filter_by(machine_id=machine.id)
        
        # Apply date range filter (prioritize over year/month)
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                query = query.filter(CTPProblemLog.problem_date >= date_from_obj)
                print(f"Applied date_from filter: {date_from_obj}")
            except ValueError as e:
                print(f"Error parsing date_from: {date_from}, Error: {str(e)}")
                return jsonify({'success': False, 'error': f'Invalid date_from format: {date_from}. Use YYYY-MM-DD'}), 400
        
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                query = query.filter(CTPProblemLog.problem_date <= date_to_obj)
                print(f"Applied date_to filter: {date_to_obj}")
            except ValueError as e:
                print(f"Error parsing date_to: {date_to}, Error: {str(e)}")
                return jsonify({'success': False, 'error': f'Invalid date_to format: {date_to}. Use YYYY-MM-DD'}), 400
        
        # Apply other filters (only if date range not specified)
        if not date_from and not date_to:
            if year:
                query = query.filter(db.extract('year', CTPProblemLog.problem_date) == int(year))
            
            if month:
                query = query.filter(db.extract('month', CTPProblemLog.problem_date) == int(month))
        
        if technician_type:
            query = query.filter(CTPProblemLog.technician_type == technician_type)
        
        if status:
            query = query.filter(CTPProblemLog.status == status)
        
        if search:
            search_pattern = f'%{search}%'
            query = query.filter(
                db.or_(
                    CTPProblemLog.problem_description.ilike(search_pattern),
                    CTPProblemLog.solution.ilike(search_pattern),
                    CTPProblemLog.technician_name.ilike(search_pattern)
                )
            )
        
        # Order by problem date descending
        query = query.order_by(CTPProblemLog.problem_date.desc())
        
        # Execute query
        logs = query.all()
        
        if not logs:
            return jsonify({'success': False, 'error': 'Tidak ada data untuk rentang tanggal yang dipilih'}), 200
        
        # Prepare data for export
        export_data = []
        for log in logs:
            export_data.append({
                'Tanggal': log.problem_date.strftime('%d %b %Y %H:%M') if log.problem_date else '-',
                'Problem': log.problem_description or '-',
                'Solusi': log.solution or '-',
                'Teknisi': log.technician_name or '-',
                'Status': 'Selesai' if log.status == 'completed' else 'Berjalan',
                'Downtime': f"{log.downtime_hours:.1f} jam" if log.downtime_hours else '-',
                'Photo Reference': f"/impact/{log.problem_photo}" if log.problem_photo else '-'
            })
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"ctp_log_{machine_nickname}_{timestamp}"
        
        # Prepare period information for export
        period_info = {
            'date_from': date_from,
            'date_to': date_to
        }
        
        if format_type.lower() == 'pdf':
            # Generate PDF
            output = generate_pdf_export(export_data, machine, filename_base, period_info)
            content_type = 'application/pdf'
            filename = f"{filename_base}.pdf"
        else:
            # Generate Excel (default)
            output = generate_excel_export(export_data, machine, filename_base, period_info)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f"{filename_base}.xlsx"
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = content_type
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        print(f"Error in export_ctp_logs: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': f'Export failed: {str(e)}'}), 500

def generate_excel_export(data, machine, filename_base, period_info=None):
    """Generate Excel export for CTP logs"""
    try:
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "CTP Problem Logs"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='FF9500', end_color='FF9500', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Add title rows
        ws.merge_cells('A1:F1')

        # Build a more descriptive machine title using name + description instead of nickname
        machine_name = getattr(machine, 'name', None) or getattr(machine, 'nickname', '')
        machine_description = getattr(machine, 'description', '') or ''
        if machine_description:
            machine_title = f"{machine_name} {machine_description}"
        else:
            machine_title = machine_name

        ws['A1'] = f"LAPORAN PROBLEM MESIN {machine_title.upper()}"
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Format period information
        period_text = ""
        if period_info:
            date_from = period_info.get('date_from')
            date_to = period_info.get('date_to')
            if date_from and date_to:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                    period_text = f"{date_from_obj.strftime('%d %B %Y')} - {date_to_obj.strftime('%d %B %Y')}"
                except ValueError:
                    period_text = f"{datetime.now().strftime('%d %B %Y')}"
            else:
                period_text = f"{datetime.now().strftime('%d %B %Y')}"
        else:
            period_text = f"{datetime.now().strftime('%d %B %Y')}"
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Periode: {period_text}"
        ws['A2'].font = Font(name='Arial', size=11)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add empty row
        ws.append([])
        
        # Define headers
        headers = ['Tanggal', 'Problem', 'Solusi', 'Teknisi', 'Status', 'Downtime']
        
        # Add header row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for row_num, row_data in enumerate(data, 5):
            ws.append([
                row_data.get('Tanggal', ''),
                row_data.get('Problem', ''),
                row_data.get('Solusi', ''),
                row_data.get('Teknisi', ''),
                row_data.get('Status', ''),
                row_data.get('Downtime', '')
            ])
            
            # Apply borders to data cells
            for col_num in range(1, 7):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Auto-adjust column widths
        column_widths = [15, 35, 35, 15, 12, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        raise Exception(f"Error generating Excel export: {str(e)}")

def generate_pdf_export(data, machine, filename_base, period_info=None):
    """Generate PDF export for CTP logs with A4 landscape and fit-to-page scaling"""
    try:
        
        # Create PDF document with A4 landscape and smaller margins for better fit
        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            leftMargin=0.5*inch,
            rightMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        styles = getSampleStyleSheet()
        
        # Container for story elements
        story = []
        
        # Build a more descriptive machine title using name + description instead of nickname
        machine_name = getattr(machine, 'name', None) or getattr(machine, 'nickname', '')
        machine_description = getattr(machine, 'description', '') or ''
        if machine_description:
            machine_title = f"{machine_name} {machine_description}"
        else:
            machine_title = machine_name

        # Add title with smaller font, using machine.name + machine.description
        title_style = styles['Title']
        title = Paragraph(f"LAPORAN PROBLEM MESIN {machine_title.upper()}", title_style)
        story.append(title)
        story.append(Spacer(1, 8))
        
        # Format period information
        period_text = ""
        if period_info:
            date_from = period_info.get('date_from')
            date_to = period_info.get('date_to')
            if date_from and date_to:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                    period_text = f"{date_from_obj.strftime('%d %B %Y')} - {date_to_obj.strftime('%d %B %Y')}"
                except ValueError:
                    period_text = f"{datetime.now().strftime('%d %B %Y')}"
            else:
                period_text = f"{datetime.now().strftime('%d %B %Y')}"
        else:
            period_text = f"{datetime.now().strftime('%d %B %Y')}"
        
        # Add period with smaller font
        period_style = styles['Normal']
        period = Paragraph(f"Periode: {period_text}", period_style)
        story.append(period)
        story.append(Spacer(1, 6))
        
        # Prepare table data (header row first)
        table_data = [['Tanggal', 'Problem', 'Solusi', 'Teknisi', 'Status', 'Downtime']]

        # Create Paragraph styles for table cells to enable proper wrapping

        body_style = ParagraphStyle(
            'BodyCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=7,
            leading=8.5,          # slightly larger than fontSize for readability
            alignment=0,          # left
            wordWrap='CJK',       # better wrapping for long text with spaces/newlines
        )

        # Helper to normalize text: handle None and replace explicit "\n" with real newlines
        def normalize_text(value):
            if value is None:
                return ''
            text = str(value)
            # Convert literal "\n" into actual newlines (in case stored like that)
            text = text.replace('\\n', '\n')
            return text

        # Build data rows with Paragraphs for Problem and Solusi so they wrap correctly
        for row_data in data:
            tanggal = normalize_text(row_data.get('Tanggal', ''))
            problem_text = normalize_text(row_data.get('Problem', ''))
            solusi_text = normalize_text(row_data.get('Solusi', ''))
            teknisi = normalize_text(row_data.get('Teknisi', ''))
            status = normalize_text(row_data.get('Status', ''))
            downtime = normalize_text(row_data.get('Downtime', ''))

            table_data.append([
                tanggal,
                Paragraph(problem_text, body_style),
                Paragraph(solusi_text, body_style),
                teknisi,
                status,
                downtime
            ])
        
        # Calculate available width for table (A4 landscape width minus margins)
        available_width = 11.69*inch - 1.0*inch  # A4 landscape width minus margins
        
        # Create table with optimized column widths to fit page
        table = Table(table_data, repeatRows=1, colWidths=[
            1.2*inch,   # Tanggal (slightly wider)
            2.5*inch,   # Problem
            2.5*inch,   # Solusi
            1.2*inch,   # Teknisi (slightly wider)
            1.0*inch,   # Status
            1.0*inch     # Downtime
        ])
        
        # Define table style with smaller fonts for better fit
        table_style = TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF9500')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),  # Smaller header font
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),  # Smaller data font
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            
            # Word wrapping for all cells including headers
            ('WORDWRAP', (0, 0), (-1, -1), True),
            
            # Borders
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
            
            # Padding to reduce space
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2)
        ])
        
        table.setStyle(table_style)
        story.append(table)
        
        # Build PDF with onFirstPage callback to ensure proper layout
        def on_first_page(canvas, doc):
            canvas.saveState()
            # Set page size to ensure A4 landscape
            canvas.setPageSize(landscape(A4))
            canvas.restoreState()
        
        doc.build(story, onFirstPage=on_first_page)
        output.seek(0)
        
        return output
        
    except Exception as e:
        raise Exception(f"Error generating PDF export: {str(e)}")
    
@export_bp.route('/export-chemical-bon')
@login_required
def export_chemical_bon():
    try:
        # Set locale to Indonesian for date formatting
        try:
            # For systems that support UTF-8 (e.g., Linux, macOS)
            locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')
        except locale.Error:
            # Fallback for systems that might not support UTF-8 suffix (e.g., some Windows versions)
            locale.setlocale(locale.LC_TIME, 'id_ID')
        
        # Get query parameters
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        brand_filter = request.args.get('brand') # This is the original filter value

        # Convert string dates to datetime objects
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None

        # Create new workbook
        wb = openpyxl.Workbook()
        
        # Define styles
        title_font = Font(bold=True, size=24)
        subtitle_font = Font(bold=True, size=18)
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Define border style once
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

        # Determine which brands to process
        if brand_filter:
            brands_to_process = [brand_filter]
            # Use the actual brand_filter for download name if a specific brand is selected
            download_brand_suffix = f"_{brand_filter.replace(' ', '_')}" 
        else:
            # Mengambil daftar brand unik dari database
            brands_to_process = db.session.query(ChemicalBonCTP.brand).distinct().all()
            brands_to_process = [brand[0] for brand in brands_to_process]
            # If no specific brand filter, use "_all" for download name
            download_brand_suffix = "_all"


        # Create sheet for each brand
        for idx, brand in enumerate(brands_to_process):
            # Build query for current brand
            query = ChemicalBonCTP.query.filter(ChemicalBonCTP.brand == brand)
            
            if start_date and end_date:
                query = query.filter(ChemicalBonCTP.tanggal.between(start_date, end_date))
            
            records = query.order_by(ChemicalBonCTP.tanggal.asc()).all()
            
            # Only create a sheet if there are records for this brand
            if not records and len(brands_to_process) > 1: # Skip if no records and multiple brands
                continue
            elif not records and len(brands_to_process) == 1: # If only one brand chosen and no records, still create an empty sheet
                if idx == 0:
                    ws = wb.active
                    ws.title = brand
                else:
                    ws = wb.create_sheet(brand)
                # Ensure the sheet has basic headers even if empty
                ws.merge_cells('A1:K1')
                ws['A1'] = 'Laporan Chemical Bon CTP'
                ws['A1'].font = title_font
                ws['A1'].alignment = center_alignment
                ws.merge_cells('A2:K2')
                ws['A2'] = f'{brand}'
                ws['A2'].font = subtitle_font
                ws['A2'].alignment = center_alignment
                
                date_range_str = ""
                if start_date and end_date:
                    date_range_str = f'{start_date.strftime("%#d %B %Y")} s/d {end_date.strftime("%#d %B %Y")}'
                elif start_date:
                    date_range_str = f'Dari {start_date.strftime("%#d %B %Y")}'
                elif end_date:
                    date_range_str = f'Sampai {end_date.strftime("%#d %B %Y")}'
                ws.merge_cells('A3:K3')
                ws['A3'] = date_range_str
                ws['A3'].font = subtitle_font
                ws['A3'].alignment = center_alignment

                headers = ['Tanggal', 'Bon Number', 'Request Number', 'Brand', 'Item Code', 'Item Name', 
                           'Unit', 'Jumlah', 'PIC', 'Keterangan', 'Periode']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
                continue # Go to next brand


            # Create or get sheet
            if idx == 0: # For the first brand, use the active sheet
                ws = wb.active
                ws.title = brand
            else: # For subsequent brands, create a new sheet
                ws = wb.create_sheet(brand)

            # Set column widths
            column_widths = [20, 20, 20, 15, 20, 40, 10, 10, 20, 30, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Write main title and subtitle
            ws.merge_cells('A1:K1')
            ws['A1'] = 'Laporan Chemical Bon CTP'
            ws['A1'].font = title_font
            ws['A1'].alignment = center_alignment

            ws.merge_cells('A2:K2')
            ws['A2'] = f'{brand}'
            ws['A2'].font = subtitle_font
            ws['A2'].alignment = center_alignment
            
            # Use `strftime` for the date range header
            date_range_str = ""
            if start_date and end_date:
                date_range_str = f'{start_date.strftime("%#d %B %Y")} s/d {end_date.strftime("%#d %B %Y")}'
            elif start_date:
                date_range_str = f'Dari {start_date.strftime("%#d %B %Y")}'
            elif end_date:
                date_range_str = f'Sampai {end_date.strftime("%#d %B %Y")}'

            ws.merge_cells('A3:K3')
            ws['A3'] = date_range_str
            ws['A3'].font = subtitle_font
            ws['A3'].alignment = center_alignment

            # Headers row
            headers = ['Tanggal', 'Bon Number', 'Request Number', 'Brand', 'Item Code', 'Item Name', 
                       'Unit', 'Jumlah', 'PIC', 'Keterangan', 'Periode']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border # Apply border to header cells

            # Write data
            for row_idx, record in enumerate(records, 6):
                # Use `strftime` for the record date
                tanggal_str = record.tanggal.strftime('%#d %B %Y') if record.tanggal else ''
                data = [
                    tanggal_str,
                    record.bon_number,
                    record.request_number,
                    record.brand,
                    record.item_code,
                    record.item_name,
                    record.unit,
                    record.jumlah,
                    record.user.name if record.user else '',
                    record.wo_number if record.wo_number else '',  # Keterangan diisi dengan WO number
                    record.bon_periode
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Use center alignment for most columns, left for specific text columns
                    cell.alignment = Alignment(horizontal='center' if col_idx not in [5, 6, 9, 10] else 'left')
                    cell.border = thin_border # Apply border to data cells

        # Remove the default empty sheet created by openpyxl if it exists and we've created other sheets
        # Or if no data was found and the default sheet is the only one
        if "Sheet" in wb.sheetnames:
            if len(brands_to_process) > 0 and len(wb.sheetnames) > 1:
                del wb["Sheet"]
            elif len(brands_to_process) == 0: # If no brands were found at all
                del wb["Sheet"]


        # Create response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Use the determined download_brand_suffix here
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'chemical_bon{download_brand_suffix}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        print(f"Error in export: {str(e)}")
        return jsonify({'success': False, 'message': f'Export gagal: {str(e)}'}), 500

@export_bp.route('/export-pdnd-adjustment', methods=['GET'])
def export_pdnd_adjustment():
    try:
        
        # Get filter parameters
        status_filter = request.args.get('status', '')
        remarks_filter = request.args.get('remarks', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Build query
        query = PlateAdjustmentRequest.query
        
        # Apply filters
        if status_filter:
            if status_filter == 'selesai':
                query = query.filter(PlateAdjustmentRequest.status.in_(['selesai', 'menunggu_adjustment_pdnd']))
            else:
                query = query.filter(PlateAdjustmentRequest.status == status_filter)
        
        if remarks_filter:
            query = query.filter(PlateAdjustmentRequest.remarks.ilike(f'%{remarks_filter}%'))
        
        if date_from:
            query = query.filter(PlateAdjustmentRequest.tanggal >= date_from)
        if date_to:
            query = query.filter(PlateAdjustmentRequest.tanggal <= date_to)
        
        # Get data sorted by date (newest first)
        adjustments = query.order_by(PlateAdjustmentRequest.tanggal.desc(), PlateAdjustmentRequest.id.desc()).all()
        
        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "PDND Adjustment Data"
        
        # Write headers
        headers = [
            'ID', 'Tanggal', 'Mesin Cetak', 'PIC', 'Remarks', 'WO Number', 'MC Number', 
            'Run Length', 'Item Name', 'Jumlah Plate', 'Note', 'Machine Off At',
            'PDND Start At', 'PDND Finish At', 'PDND By',
            'Mounting Start At', 'Mounting Finish At', 'Mounting By',
            'Plate Start At', 'Plate Finish At', 'Plate Delivered At', 'CTP By', 'Status'
        ]
        worksheet.append(headers)
        
        # Write data rows
        for adj in adjustments:
            row_data = [
                str(adj.id or ''),
                adj.tanggal.strftime('%Y-%m-%d') if adj.tanggal else '',
                str(adj.mesin_cetak or ''),
                str(adj.pic or ''),
                str(adj.remarks or ''),
                str(adj.wo_number or ''),
                str(adj.mc_number or ''),
                str(adj.run_length or ''),
                str(adj.item_name or ''),
                str(adj.jumlah_plate or ''),
                str(adj.note or ''),
                adj.machine_off_at.strftime('%Y-%m-%d %H:%M:%S') if adj.machine_off_at else '',
                adj.pdnd_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.pdnd_start_at else '',
                adj.pdnd_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.pdnd_finish_at else '',
                str(adj.pdnd_by or ''),
                adj.adjustment_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_start_at else '',
                adj.adjustment_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_finish_at else '',
                str(adj.adjustment_by or ''),
                adj.plate_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_start_at else '',
                adj.plate_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_finish_at else '',
                adj.plate_delivered_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_delivered_at else '',
                str(adj.ctp_by or ''),
                str(adj.status or '')
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=pdnd_adjustment_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exporting data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@export_bp.route('/export-curve-adjustment', methods=['GET'])
def export_curve_adjustment():
    try:
        
        # Get filter parameters
        status_filter = request.args.get('status', '')
        remarks_filter = request.args.get('remarks', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Build query
        query = PlateAdjustmentRequest.query
        
        # Apply filters
        if status_filter:
            if status_filter == 'selesai':
                query = query.filter(PlateAdjustmentRequest.status.in_(['selesai', 'menunggu_adjustment_curve']))
            else:
                query = query.filter(PlateAdjustmentRequest.status == status_filter)
        
        if remarks_filter:
            query = query.filter(PlateAdjustmentRequest.remarks.ilike(f'%{remarks_filter}%'))
        
        if date_from:
            query = query.filter(PlateAdjustmentRequest.tanggal >= date_from)
        if date_to:
            query = query.filter(PlateAdjustmentRequest.tanggal <= date_to)
        
        # Get data sorted by date (newest first)
        adjustments = query.order_by(PlateAdjustmentRequest.tanggal.desc(), PlateAdjustmentRequest.id.desc()).all()
        
        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Curve Adjustment Data"
        
        # Write headers
        headers = [
            'ID', 'Tanggal', 'Mesin Cetak', 'PIC', 'Remarks', 'WO Number', 'MC Number', 
            'Run Length', 'Item Name', 'Jumlah Plate', 'Note', 'Machine Off At',
            'Curve Start At', 'Curve Finish At', 'Curve By',
            'Mounting Start At', 'Mounting Finish At', 'Mounting By', 
            'Plate Start At', 'Plate Finish At', 'Plate Delivered At', 'CTP By', 'Status'
        ]
        worksheet.append(headers)
        
        # Write data rows
        for adj in adjustments:
            row_data = [
                str(adj.id or ''),
                adj.tanggal.strftime('%Y-%m-%d') if adj.tanggal else '',
                str(adj.mesin_cetak or ''),
                str(adj.pic or ''),
                str(adj.remarks or ''),
                str(adj.wo_number or ''),
                str(adj.mc_number or ''),
                str(adj.run_length or ''),
                str(adj.item_name or ''),
                str(adj.jumlah_plate or ''),
                str(adj.note or ''),
                adj.machine_off_at.strftime('%Y-%m-%d %H:%M:%S') if adj.machine_off_at else '',
                adj.curve_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.curve_start_at else '',
                adj.curve_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.curve_finish_at else '',
                str(adj.curve_by or ''),
                adj.adjustment_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_start_at else '',
                adj.adjustment_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_finish_at else '',
                str(adj.adjustment_by or ''),
                adj.plate_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_start_at else '',
                adj.plate_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_finish_at else '',
                adj.plate_delivered_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_delivered_at else '',
                str(adj.ctp_by or ''),
                str(adj.status or '')
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=curve_adjustment_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exporting data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500 

@export_bp.route('/export-design-adjustment', methods=['GET'])
def export_design_adjustment():
    try:
        
        # Get filter parameters
        status_filter = request.args.get('status', '')
        remarks_filter = request.args.get('remarks', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Build query
        query = PlateAdjustmentRequest.query
        
        # Apply filters
        if status_filter:
            if status_filter == 'selesai':
                query = query.filter(PlateAdjustmentRequest.status.in_(['selesai', 'menunggu_adjustment_design']))
            else:
                query = query.filter(PlateAdjustmentRequest.status == status_filter)
        
        if remarks_filter:
            query = query.filter(PlateAdjustmentRequest.remarks.ilike(f'%{remarks_filter}%'))
        
        if date_from:
            query = query.filter(PlateAdjustmentRequest.tanggal >= date_from)
        if date_to:
            query = query.filter(PlateAdjustmentRequest.tanggal <= date_to)
        
        # Get data sorted by date (newest first)
        adjustments = query.order_by(PlateAdjustmentRequest.tanggal.desc(), PlateAdjustmentRequest.id.desc()).all()
        
        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Design Adjustment Data"
        
        # Write headers
        headers = [
            'ID', 'Tanggal', 'Mesin Cetak', 'PIC', 'Remarks', 'WO Number', 'MC Number', 
            'Run Length', 'Item Name', 'Jumlah Plate', 'Note', 'Machine Off At',
            'Design Start At', 'Design Finish At', 'Design By',
            'Adjustment Start At', 'Adjustment Finish At', 'Adjustment By', 
            'Plate Start At', 'Plate Finish At', 'Plate Delivered At', 'CTP By', 'Status'
        ]
        worksheet.append(headers)
        
        # Write data rows
        for adj in adjustments:
            row_data = [
                str(adj.id or ''),
                adj.tanggal.strftime('%Y-%m-%d') if adj.tanggal else '',
                str(adj.mesin_cetak or ''),
                str(adj.pic or ''),
                str(adj.remarks or ''),
                str(adj.wo_number or ''),
                str(adj.mc_number or ''),
                str(adj.run_length or ''),
                str(adj.item_name or ''),
                str(adj.jumlah_plate or ''),
                str(adj.note or ''),
                adj.machine_off_at.strftime('%Y-%m-%d %H:%M:%S') if adj.machine_off_at else '',
                adj.design_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.design_start_at else '',
                adj.design_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.design_finish_at else '',
                str(adj.design_by or ''),
                adj.adjustment_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_start_at else '',
                adj.adjustment_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_finish_at else '',
                str(adj.adjustment_by or ''),
                adj.plate_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_start_at else '',
                adj.plate_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_finish_at else '',
                adj.plate_delivered_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_delivered_at else '',
                str(adj.ctp_by or ''),
                str(adj.status or '')
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=design_adjustment_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exporting data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
@export_bp.route('/export-stock-opname')
@login_required
def export_stock_opname():
    try:
        # Set locale to Indonesian for date formatting
        try:
            locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')
        except locale.Error:
            try:
                locale.setlocale(locale.LC_TIME, 'id_ID')
            except locale.Error:
                # Fallback if specific locale not found, often on Windows systems
                locale.setlocale(locale.LC_TIME, 'Indonesian_Indonesia.1252') # Common for Windows
                
        # Get filter parameters
        date_from_str = request.args.get('date_from', '')
        date_to_str = request.args.get('date_to', '')
        jenis_plate = request.args.get('jenis_plate', '')

        # Parse dates and format them for display in headers
        date_from_formatted = ""
        date_to_formatted = ""
        if date_from_str:
            date_from_obj = datetime.strptime(date_from_str, '%Y-%m-%d')
            date_from_formatted = date_from_obj.strftime('%d %B %Y')
        if date_to_str:
            date_to_obj = datetime.strptime(date_to_str, '%Y-%m-%d')
            date_to_formatted = date_to_obj.strftime('%d %B %Y')

        # Mapping data untuk setiap jenis plate (Didefinisikan di dalam route)
        plate_details = {
            'SAPHIRA 1030': {
                'size': '1030 X 790 MM',
                'item_code': '02-049-000-0000002',
                'item_name': '(SUT1.PAO1SX1) SAPHIRA PA.27 27x1030x790 PKT50 (BOX 50PCS)'
            },
            'SAPHIRA 1030 PN': {
                'size': '1030 X 790 MM',
                'item_code': '02-023-000-0000006',
                'item_name': '(SUT1.PNO7UWO) SAPHIRA PN 30 1030 X 790 MM PKT40 (BOX 40PCS)'
            },            
            'SAPHIRA 1055': {
                'size': '1055 X 811 MM',
                'item_code': '02-049-000-0000003',
                'item_name': '(SUT1.PAO1SY3) SAPHIRA PA.27 27X1055X811 PKT50 (BOX 50PCS)'
            },
            'SAPHIRA 1055 PN': {
                'size': '1055 X 811 MM',
                'item_code': '02-049-000-0000011',
                'item_name': '(SUT1.PNO7U8C) SAPHIRA PN 30 1055 X 811 MM PKT40 (BOX 40PCS)'
            },
            'SAPHIRA 1630': {
                'size': '1630 X 1325 MM',
                'item_code': '02-049-000-0000001',
                'item_name': '(SUT1.PNOQXG8) SAPHIRA PN 40 1630 1325 PKT 30 (BOX 30PCS)'
            },
            'FUJI 1030': {
                'size': '1030 X 790 MM',
                'item_code': '02-049-000-0000008',
                'item_name': 'PLATE FUJI LH-PK 1030x790x0.3 (BOX 30PCS)'
            },
            'FUJI 1030 LHPJA': {
                'size': '1030 X 790 MM',
                'item_code': '02-049-000-0000012',
                'item_name': 'PLATE FUJI LH-PJA 1030x790x0.3 (BOX 30PCS)'
            },            
            'FUJI 1030 UV': {
                'size': '1030 X 790 MM',
                'item_code': '02-023-000-0000007',
                'item_name': 'PLATE FUJI LH-PJ2 1030x790x0.3 (BOX 30PCS)'
            },
            'FUJI 1055': {
                'size': '1055 X 811 MM',
                'item_code': '02-049-000-0000010',
                'item_name': 'PLATE FUJI LH-PK 1055x811x0.3 (BOX 30PCS)'
            },
            'FUJI 1055 LHPL': {
                'size': '1055 X 811 MM',
                'item_code': '02-049-000-0000013',
                'item_name': 'PLATE FUJI LH-PL 1055x811x0.3 (BOX 30PCS)'
            },
            'FUJI 1055 UV': {
                'size': '1055 X 811 MM',
                'item_code': '02-023-000-0000012',
                'item_name': 'PLATE FUJI LH-PJ2 1055x811x0.3 (BOX 30PCS)'
            },
            'FUJI 1630': {
                'size': '1630 X 1325 MM',
                'item_code': '02-049-000-0000009',
                'item_name': 'PLATE FUJI LH-PJ2 1630x1325x0.4 (BOX 15PCS)'
            }
        }

        # Define styles (Didefinisikan di dalam route)
        title_font = Font(bold=True, size=24)
        subtitle_font = Font(bold=True, size=18)
        info_font = Font(bold=True, size=11)
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        total_row_fill_green = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')
        total_row_fill_grey = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

        wb = Workbook()

        # Helper function untuk menambahkan baris "Total Pemakaian" (Didefinisikan di dalam route)
        def insert_total_pemakaian_row_local(ws, row_num, total_qty_ok, total_qty_ng):
            # Merge A-E for "Total Pemakaian"
            ws.merge_cells(start_row=row_num, end_row=row_num, start_column=1, end_column=5)
            total_text_cell = ws.cell(row=row_num, column=1, value="Total Pemakaian")
            total_text_cell.font = Font(bold=True)
            total_text_cell.alignment = center_alignment 
            total_text_cell.border = border
            total_text_cell.fill = total_row_fill_green

            # Merge F-H for total quantity
            ws.merge_cells(start_row=row_num, end_row=row_num, start_column=6, end_column=8)
            qty_total_cell = ws.cell(row=row_num, column=6, value=total_qty_ok + total_qty_ng)
            qty_total_cell.font = Font(bold=True)
            qty_total_cell.alignment = center_alignment 
            qty_total_cell.border = border
            qty_total_cell.fill = total_row_fill_green

            # Merge I-J for border fill (grey)
            ws.merge_cells(start_row=row_num, end_row=row_num, start_column=9, end_column=10)
            fill_cell_i = ws.cell(row=row_num, column=9) 
            fill_cell_i.border = border
            fill_cell_i.fill = total_row_fill_grey
            
            # Ensure the last column of the merged range also has a border, though fill is applied to the first cell
            fill_cell_j = ws.cell(row=row_num, column=10)
            fill_cell_j.border = border

        # Helper function untuk menulis data ke sheet (Didefinisikan di dalam route)
        def write_sheet_data_local(ws, records, plate_type_current, date_from_formatted, date_to_formatted):
            # Add header information
            ws.cell(row=1, column=1, value="Laporan Stock Opname").font = title_font
            ws.merge_cells('A1:J1')
            ws.cell(row=1, column=1).alignment = center_alignment

            ws.cell(row=2, column=1, value=plate_type_current).font = subtitle_font
            ws.merge_cells('A2:J2')
            ws.cell(row=2, column=1).alignment = center_alignment

            date_range_text = f"{date_from_formatted} - {date_to_formatted}"
            ws.cell(row=3, column=1, value=date_range_text).font = subtitle_font
            ws.merge_cells('A3:J3')
            ws.cell(row=3, column=1).alignment = center_alignment

            details = plate_details.get(plate_type_current, {})
            ws.cell(row=4, column=1, value="Size").font = info_font
            ws.cell(row=4, column=2, value=f": {details.get('size', '')}").font = info_font
            ws.cell(row=5, column=1, value="Item Code").font = info_font
            ws.cell(row=5, column=2, value=f": {details.get('item_code', '')}").font = info_font
            ws.cell(row=6, column=1, value="Item Name").font = info_font
            ws.cell(row=6, column=2, value=f": {details.get('item_name', '')}").font = info_font

            # Write headers on row 8
            headers = [
                'Tanggal', 'No WO', 'Mesin CTP', 'Mesin Cetak', 'Nama Item',
                'Qty OK', 'Qty NG', 'Total', 'Keterangan Not Good', 'Catatan'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=8, column=col)
                cell.value = header
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
                cell.fill = header_fill

            current_row = 9
            
            if not records:
                ws.cell(row=current_row, column=1, value="Tidak ada data untuk periode ini.")
                ws.merge_cells(f"A{current_row}:J{current_row}")
                ws.cell(row=current_row, column=1).alignment = center_alignment
                # Apply border to the merged cell
                for col_idx in range(1, 11):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.border = border
                return # Exit if no records

            current_date = None
            date_group_start_row = 0
            date_group_total_qty_ok = 0
            date_group_total_qty_ng = 0

            for record_idx, record in enumerate(records):
                record_formatted_date = record.log_date.strftime('%d %B %Y')

                # Check if it's a new date group
                if record_formatted_date != current_date:
                    # If not the first group, finalize the previous group (merge date & add total row)
                    if current_date is not None:
                        # Merge date column for the previous group
                        ws.merge_cells(start_row=date_group_start_row, end_row=current_row - 1, start_column=1, end_column=1)
                        ws.cell(row=date_group_start_row, column=1).alignment = center_alignment
                        
                        # Add "Total Pemakaian" row for the previous group
                        insert_total_pemakaian_row_local(
                            ws, current_row, date_group_total_qty_ok, date_group_total_qty_ng
                        )
                        current_row += 1 # Increment for the total row

                    # Start a new date group
                    current_date = record_formatted_date
                    date_group_start_row = current_row
                    date_group_total_qty_ok = 0
                    date_group_total_qty_ng = 0
                
                # Process current record
                wo_numbers = record.wo_number.split(',') if record.wo_number else ['']
                wo_numbers = [wo.strip() for wo in wo_numbers]
                
                qty_ok_record = record.num_plate_good or 0
                qty_ng_record = record.num_plate_not_good or 0
                total_record_qty = qty_ok_record + qty_ng_record

                # Accumulate totals for the current date group
                date_group_total_qty_ok += qty_ok_record
                date_group_total_qty_ng += qty_ng_record

                # Define columns that should be merged if num_rows > 1 (excluding Tanggal (1) and No WO (2))
                record_specific_merge_cols = [3, 4, 5, 6, 7, 8, 9, 10]

                if len(wo_numbers) > 1:
                    # Perform merge for these columns for this record's multiple WO lines
                    for col_idx in record_specific_merge_cols:
                        ws.merge_cells(
                            start_row=current_row,
                            start_column=col_idx,
                            end_row=current_row + len(wo_numbers) - 1,
                            end_column=col_idx
                        )
                
                # Write data for the first line of this record (including the first WO)
                initial_row_values = [
                    record_formatted_date, # This value is explicitly set below if it's the start of a date group
                    wo_numbers[0],
                    record.ctp_machine,
                    record.print_machine,
                    record.item_name,
                    qty_ok_record,
                    qty_ng_record,
                    total_record_qty,
                    record.not_good_reason or '',
                    record.note or ''
                ]

                for col_idx, value in enumerate(initial_row_values, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    # For Tanggal (Column 1): only set value if it's the start of a date group.
                    # Otherwise, it's part of a merged cell, just apply styling.
                    if col_idx == 1:
                        if current_row == date_group_start_row:
                            cell.value = value
                        # Else: no value for merged cells (it inherits from the top-left merged cell)
                    # For other columns (No WO, Mesin CTP, etc.): set value for the first line of the record.
                    # If len(wo_numbers) > 1, these columns will be merged, so value only goes in the first row.
                    else: 
                        cell.value = value
                    
                    cell.alignment = center_alignment 
                    cell.border = border


                # Handle additional WO numbers for the same record (if wo_number is split)
                for i, wo in enumerate(wo_numbers[1:], 1):
                    current_row += 1
                    
                    # Column 1 (Tanggal): This cell is part of a date merged range. Only apply styling.
                    cell = ws.cell(row=current_row, column=1)
                    cell.alignment = center_alignment
                    cell.border = border

                    # Column 2 (No WO): Gets its specific WO number.
                    wo_cell = ws.cell(row=current_row, column=2)
                    wo_cell.value = wo
                    wo_cell.alignment = center_alignment
                    wo_cell.border = border
                    
                    # For other columns (3-10): these are merged for this record. Only apply styling.
                    for col_idx in record_specific_merge_cols:
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.alignment = center_alignment
                        cell.border = border
                    
                current_row += 1 # Move to the next row after all WO numbers for this record

            # After the loop, finalize the very last date group
            if current_date is not None:
                ws.merge_cells(start_row=date_group_start_row, end_row=current_row - 1, start_column=1, end_column=1)
                ws.cell(row=date_group_start_row, column=1).alignment = center_alignment
                
                insert_total_pemakaian_row_local(
                    ws, current_row, date_group_total_qty_ok, date_group_total_qty_ng
                )
                current_row += 1 # Increment for the total row

            # Set column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15


        if not jenis_plate:
            # Skenario: Semua Jenis Plate (banyak sheet)
            unique_plates = db.session.query(CTPProductionLog.plate_type_material).distinct().all()
            unique_plates = [p[0] for p in unique_plates if p[0] and p[0] in plate_details]
            
            # Mengurutkan jenis plate berdasarkan abjad
            unique_plates.sort()

            default_ws = wb.active
            wb.remove(default_ws)

            for plate_type in unique_plates:
                ws = wb.create_sheet(title=plate_type[:31]) # Max 31 chars for sheet title

                query = db.session.query(CTPProductionLog).filter(CTPProductionLog.plate_type_material == plate_type)
                if date_from_str:
                    query = query.filter(CTPProductionLog.log_date >= date_from_str)
                if date_to_str:
                    date_to_obj_end = datetime.strptime(date_to_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                    query = query.filter(CTPProductionLog.log_date <= date_to_obj_end)
                
                records = query.order_by(CTPProductionLog.log_date.desc(), CTPProductionLog.id.desc()).all()
                
                write_sheet_data_local(ws, records, plate_type, date_from_formatted, date_to_formatted)

        else:
            # Skenario: Satu Jenis Plate (satu sheet)
            ws = wb.active
            ws.title = jenis_plate[:31]

            query = db.session.query(CTPProductionLog).filter(CTPProductionLog.plate_type_material == jenis_plate)
            if date_from_str:
                query = query.filter(CTPProductionLog.log_date >= date_from_str)
            if date_to_str:
                date_to_obj_end = datetime.strptime(date_to_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                query = query.filter(CTPProductionLog.log_date <= date_to_obj_end)
            
            records = query.order_by(CTPProductionLog.log_date.desc(), CTPProductionLog.id.desc()).all()

            write_sheet_data_local(ws, records, jenis_plate, date_from_formatted, date_to_formatted)


        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Nama file
        jenis_plate_str = f"_{jenis_plate.replace(' ', '_')}" if jenis_plate else "_all"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'stock_opname_ctp{jenis_plate_str}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        print(f"Error exporting stock opname data: {e}")
        traceback.print_exc() 
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@export_bp.route('/export-ctp-adjustment', methods=['GET'])
def export_ctp_adjustment():
    try:

        # Get filter parameters
        status_filter = request.args.get('status', '')
        mesin_filter = request.args.get('mesin', '')
        remarks_filter = request.args.get('remarks', '')
        search_filter = request.args.get('search', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')

        # Build query for CTP adjustments
        query = PlateAdjustmentRequest.query

        # Apply filters
        if status_filter:
            if status_filter == 'selesai':
                query = query.filter(PlateAdjustmentRequest.status.in_(['selesai', 'proses_ctp']))
            else:
                query = query.filter(PlateAdjustmentRequest.status == status_filter)

        if mesin_filter:
            query = query.filter(PlateAdjustmentRequest.mesin_cetak == mesin_filter)
            
        if remarks_filter:
            query = query.filter(PlateAdjustmentRequest.remarks.ilike(f'%{remarks_filter}%'))
            
        if search_filter:
            search_term = f'%{search_filter}%'
            query = query.filter(
                db.or_(
                    PlateAdjustmentRequest.wo_number.ilike(search_term),
                    PlateAdjustmentRequest.mc_number.ilike(search_term),
                    PlateAdjustmentRequest.item_name.ilike(search_term),
                    PlateAdjustmentRequest.mesin_cetak.ilike(search_term),
                    PlateAdjustmentRequest.remarks.ilike(search_term)
                )
            )
        
        if date_from:
            query = query.filter(PlateAdjustmentRequest.tanggal >= date_from)
        if date_to:
            query = query.filter(PlateAdjustmentRequest.tanggal <= date_to)
        
        # Get CTP data sorted by date (newest first)
        adjustments = query.order_by(PlateAdjustmentRequest.tanggal.desc(), PlateAdjustmentRequest.id.desc()).all()
        
        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "CTP Adjustment Data"

        # Write headers
        headers = [
            'ID', 'Tanggal', 'Mesin Cetak', 'PIC', 'Remarks', 'WO Number', 'MC Number', 
            'Run Length', 'Item Name', 'Jumlah Plate', 'Note', 'Machine Off At',
            'PDND Start At', 'PDND Finish At', 'PDND By',
            'Design Start At', 'Design Finish At', 'Design By',
            'Curve Start At', 'Curve Finish At', 'Curve By',            
            'Adjustment Start At', 'Adjustment Finish At', 'Adjustment By', 
            'Plate Start At', 'Plate Finish At', 'Plate Delivered At', 'CTP By', 'Status'
        ]
        worksheet.append(headers)

        # Write data rows
        for adj in adjustments:
            row_data = [
                str(adj.id or ''),
                adj.tanggal.strftime('%Y-%m-%d') if adj.tanggal else '',
                str(adj.mesin_cetak or ''),
                str(adj.pic or ''),
                str(adj.remarks or ''),
                str(adj.wo_number or ''),
                str(adj.mc_number or ''),
                str(adj.run_length or ''),
                str(adj.item_name or ''),
                str(adj.jumlah_plate or ''),
                str(adj.note or ''),
                adj.machine_off_at.strftime('%Y-%m-%d %H:%M:%S') if adj.machine_off_at else '',
                adj.pdnd_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.pdnd_start_at else '',
                adj.pdnd_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.pdnd_finish_at else '',
                str(adj.pdnd_by or ''),
                adj.design_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.design_start_at else '',
                adj.design_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.design_finish_at else '',
                str(adj.design_by or ''),
                adj.curve_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.curve_start_at else '',
                adj.curve_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.curve_finish_at else '',
                str(adj.curve_by or ''),                
                adj.adjustment_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_start_at else '',
                adj.adjustment_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_finish_at else '',
                str(adj.adjustment_by or ''),
                adj.plate_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_start_at else '',
                adj.plate_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_finish_at else '',
                adj.plate_delivered_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_delivered_at else '',
                str(adj.ctp_by or ''),
                str(adj.status or '')
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=ctp_adjustment_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exporting CTP data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@export_bp.route('/export-mounting-adjustment', methods=['GET'])
def export_mounting_adjustment():
    try:
        
        # Get filter parameters
        status_filter = request.args.get('status', '')
        remarks_filter = request.args.get('remarks', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Build query
        query = PlateAdjustmentRequest.query
        
        # Apply filters
        if status_filter:
            if status_filter == 'selesai':
                query = query.filter(PlateAdjustmentRequest.status.in_(['selesai', 'proses_ctp']))
            else:
                query = query.filter(PlateAdjustmentRequest.status == status_filter)
        
        if remarks_filter:
            query = query.filter(PlateAdjustmentRequest.remarks.ilike(f'%{remarks_filter}%'))
        
        if date_from:
            query = query.filter(PlateAdjustmentRequest.tanggal >= date_from)
        if date_to:
            query = query.filter(PlateAdjustmentRequest.tanggal <= date_to)
        
        # Get data sorted by date (newest first)
        adjustments = query.order_by(PlateAdjustmentRequest.tanggal.desc(), PlateAdjustmentRequest.id.desc()).all()
        
        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Mounting Adjustment Data"
        
        # Write headers
        headers = [
            'ID', 'Tanggal', 'Mesin Cetak', 'PIC', 'Remarks', 'WO Number', 'MC Number', 
            'Run Length', 'Item Name', 'Jumlah Plate', 'Note', 'Machine Off At',
            'PDND Start At', 'PDND Finish At', 'PDND By',
            'Design Start At', 'Design Finish At', 'Design By',
            'Curve Start At', 'Curve Finish At', 'Curve By',
            'Adjustment Start At', 'Adjustment Finish At', 'Adjustment By', 
            'Plate Start At', 'Plate Finish At', 'Plate Delivered At', 'CTP By', 'Status'
        ]
        worksheet.append(headers)
        
        # Write data rows
        for adj in adjustments:
            row_data = [
                str(adj.id or ''),
                adj.tanggal.strftime('%Y-%m-%d') if adj.tanggal else '',
                str(adj.mesin_cetak or ''),
                str(adj.pic or ''),
                str(adj.remarks or ''),
                str(adj.wo_number or ''),
                str(adj.mc_number or ''),
                str(adj.run_length or ''),
                str(adj.item_name or ''),
                str(adj.jumlah_plate or ''),
                str(adj.note or ''),
                adj.machine_off_at.strftime('%Y-%m-%d %H:%M:%S') if adj.machine_off_at else '',
                adj.pdnd_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.pdnd_start_at else '',
                adj.pdnd_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.pdnd_finish_at else '',
                str(adj.pdnd_by or ''),
                adj.design_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.design_start_at else '',
                adj.design_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.design_finish_at else '',
                str(adj.design_by or ''),
                adj.curve_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.curve_start_at else '',
                adj.curve_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.curve_finish_at else '',
                str(adj.curve_by or ''),
                adj.adjustment_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_start_at else '',
                adj.adjustment_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_finish_at else '',
                str(adj.adjustment_by or ''),
                adj.plate_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_start_at else '',
                adj.plate_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_finish_at else '',
                adj.plate_delivered_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_delivered_at else '',
                str(adj.ctp_by or ''),
                str(adj.status or '')
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=mounting_adjustment_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exporting data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
@export_bp.route('/export-ctp-adjustment-data', methods=['GET'])
def export_ctp_adjustment_data():
    try:
        
        # Get filter parameters
        status_filter = request.args.get('status', '')
        remarks_filter = request.args.get('remarks', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Define the filename with .xlsx extension
        filename = f'ctp_adjustment_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Build query
        query = PlateAdjustmentRequest.query
        
        # Apply filters
        if status_filter:
            if status_filter == 'selesai':
                query = query.filter(PlateAdjustmentRequest.status == 'selesai')
            else:
                query = query.filter(PlateAdjustmentRequest.status == status_filter)
        
        if remarks_filter:
            query = query.filter(PlateAdjustmentRequest.remarks.ilike(f'%{remarks_filter}%'))
            
        if start_date:
            query = query.filter(PlateAdjustmentRequest.tanggal >= start_date)
        if end_date:
            query = query.filter(PlateAdjustmentRequest.tanggal <= end_date)
        
        # Get data sorted by date (newest first)
        adjustment_requests = query.order_by(PlateAdjustmentRequest.tanggal.desc(), PlateAdjustmentRequest.id.desc()).all()
        
        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "CTP Adjustment Data"
        
        # Write headers
        headers = [
            'ID', 'Tanggal', 'Mesin Cetak', 'PIC', 'Remarks', 'WO Number', 'MC Number', 
            'Run Length', 'Item Name', 'Jumlah Plate', 'Note', 'Machine Off At',
            'PDND Start At', 'PDND Finish At', 'PDND By',
            'Design Start At', 'Design Finish At', 'Design By',
            'Curve Start At', 'Curve Finish At', 'Curve By',            
            'Adjustment Start At', 'Adjustment Finish At', 'Adjustment By', 
            'Plate Start At', 'Plate Finish At', 'Plate Delivered At', 'CTP By', 'Status'
        ]
        worksheet.append(headers)
        
        # Write data rows
        for adj in adjustment_requests:
            row_data = [
                str(adj.id or ''),
                adj.tanggal.strftime('%Y-%m-%d') if adj.tanggal else '',
                str(adj.mesin_cetak or ''),
                str(adj.pic or ''),
                str(adj.remarks or ''),
                str(adj.wo_number or ''),
                str(adj.mc_number or ''),
                str(adj.run_length or ''),
                str(adj.item_name or ''),
                str(adj.jumlah_plate or ''),
                str(adj.note or ''),
                adj.machine_off_at.strftime('%Y-%m-%d %H:%M:%S') if adj.machine_off_at else '',
                adj.pdnd_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.pdnd_start_at else '',
                adj.pdnd_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.pdnd_finish_at else '',
                str(adj.pdnd_by or ''),
                adj.design_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.design_start_at else '',
                adj.design_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.design_finish_at else '',
                str(adj.design_by or ''),
                adj.curve_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.curve_start_at else '',
                adj.curve_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.curve_finish_at else '',
                str(adj.curve_by or ''),
                adj.adjustment_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_start_at else '',
                adj.adjustment_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.adjustment_finish_at else '',
                str(adj.adjustment_by or ''),
                adj.plate_start_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_start_at else '',
                adj.plate_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_finish_at else '',
                adj.plate_delivered_at.strftime('%Y-%m-%d %H:%M:%S') if adj.plate_delivered_at else '',
                str(adj.ctp_by or ''),
                str(adj.status or '')
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        print(f"Error exporting CTP adjustment data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@export_bp.route('/export-adjustment-press', methods=['GET'])
def export_adjustment_press():
    try:
        
        # Get filter parameters
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        mesin_cetak = request.args.get('mesin_cetak', '')

        # Build query for Adjustment Press
        query = PlateAdjustmentRequest.query
        
        # Apply filters
        if date_from:
            query = query.filter(PlateAdjustmentRequest.tanggal >= date_from)
        if date_to:
            query = query.filter(PlateAdjustmentRequest.tanggal <= date_to)
        if mesin_cetak:
            query = query.filter(PlateAdjustmentRequest.mesin_cetak == mesin_cetak)

        # Get Adjustment Press sorted by date (newest first)
        adjustment_data = query.order_by(PlateAdjustmentRequest.tanggal.desc(), PlateAdjustmentRequest.id.desc()).all()

        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Adjustment Press Data"
        
        # Write headers
        headers = [
            'ID', 'Tanggal', 'PIC', 'Mesin Cetak', 'Remarks', 'Nomor WO', 'Nomor MC',
            'Run Length', 'Nama Item', 'Jumlah Plate', 'Note', 'Mesin Off',
            'PDND Start', 'PDND Selesai', 'PIC PDND',
            'Design Start', 'Design Selesai', 'Design By',
            'Curve Start', 'Curve Selesai', 'Curve By',            
            'Adjustment Start', 'Adjustment Selesai', 'PIC Adjustment',
            'Plate Start', 'Plate Selesai', 'Plate Sampai',
            'PIC CTP', 'Grup CTP', 'Status', 'Total Downtime (jam)'
        ]
        worksheet.append(headers)
        
        # Write data rows
        for adjustment in adjustment_data:
            # Konversi string 'machine_off_at' ke objek datetime jika tidak kosong
            machine_off_dt = None
            if adjustment.machine_off_at:
                try:
                    # Asumsikan format string adalah 'YYYY-MM-DD HH:MM:SS'
                    machine_off_dt = datetime.strptime(str(adjustment.machine_off_at), '%Y-%m-%d %H:%M:%S')
                except (ValueError, TypeError):
                    # Atau format lain, coba sesuaikan
                    try:
                        machine_off_dt = datetime.strptime(str(adjustment.machine_off_at), '%d/%m/%Y %H:%M')
                    except (ValueError, TypeError):
                        pass # Biarkan machine_off_dt tetap None jika konversi gagal

            # Hitung downtime hanya jika kedua variabel adalah objek datetime yang valid
            total_downtime_hours = ''
            if adjustment.plate_delivered_at and machine_off_dt:
                time_delta = adjustment.plate_delivered_at - machine_off_dt
                total_downtime_hours = round(time_delta.total_seconds() / 3600, 2) # Dibulatkan 2 desimal

            row_data = [
                str(adjustment.id or ''),
                adjustment.tanggal.strftime('%Y-%m-%d') if adjustment.tanggal else '',
                str(adjustment.pic or ''),
                str(adjustment.mesin_cetak or ''),
                str(adjustment.remarks or ''),
                str(adjustment.wo_number or ''),
                str(adjustment.mc_number or ''),
                str(adjustment.run_length or ''),
                str(adjustment.item_name or ''),
                str(adjustment.jumlah_plate or ''),
                str(adjustment.note or ''),
                str(adjustment.machine_off_at or ''),
                adjustment.pdnd_start_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.pdnd_start_at else '',
                adjustment.pdnd_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.pdnd_finish_at else '',
                str(adjustment.pdnd_by or ''),
                adjustment.design_start_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.design_start_at else '',
                adjustment.design_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.design_finish_at else '',
                str(adjustment.design_by or ''),
                adjustment.curve_start_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.curve_start_at else '',
                adjustment.curve_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.curve_finish_at else '',
                str(adjustment.curve_by or ''),                
                adjustment.adjustment_start_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.adjustment_start_at else '',
                adjustment.adjustment_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.adjustment_finish_at else '',
                str(adjustment.adjustment_by or ''),
                adjustment.plate_start_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.plate_start_at else '',
                adjustment.plate_finish_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.plate_finish_at else '',
                adjustment.plate_delivered_at.strftime('%Y-%m-%d %H:%M:%S') if adjustment.plate_delivered_at else '',
                str(adjustment.ctp_by or ''),
                str(adjustment.ctp_group or ''),
                str(adjustment.status or ''),
                total_downtime_hours
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=adjustment_press_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exporting adjustment press data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@export_bp.route('/export-bon-press', methods=['GET'])
def export_bon_press():
    try:
        
        # Get filter parameters
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        mesin_cetak = request.args.get('mesin_cetak', '')

        # Build query for Bon Press
        query = PlateBonRequest.query
        
        # Apply filters
        if date_from:
            query = query.filter(PlateBonRequest.tanggal >= date_from)
        if date_to:
            query = query.filter(PlateBonRequest.tanggal <= date_to)
        if mesin_cetak:
            query = query.filter(PlateBonRequest.mesin_cetak == mesin_cetak)

        # Get Bon Press sorted by date (newest first)
        bon_data = query.order_by(PlateBonRequest.tanggal.desc(), PlateBonRequest.id.desc()).all()

        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Bon Press Data"
        
        # Write headers
        headers = [
            'Tanggal', 'PIC', 'Mesin Cetak', 'Remarks', 'Nomor WO', 'Nomor MC',
            'Run Length', 'Nama Item', 'Jumlah Plate', 'Note',
            'Mesin Off', 'Plate Start', 'Plate Selesai', 'Plate Sampai',
            'PIC CTP', 'Grup CTP', 'Status', 'Total Downtime (jam)'
        ]
        worksheet.append(headers)
        
        # Write data rows
        for bon in bon_data:
            # Konversi string 'machine_off_at' ke objek datetime jika tidak kosong
            machine_off_dt = None
            if bon.machine_off_at:
                try:
                    # Asumsikan format string adalah 'YYYY-MM-DD HH:MM:SS'
                    machine_off_dt = datetime.strptime(str(bon.machine_off_at), '%Y-%m-%d %H:%M:%S')
                except (ValueError, TypeError):
                    # Atau format lain jika diperlukan
                    pass

            # Hitung downtime hanya jika kedua variabel adalah objek datetime yang valid
            total_downtime_hours = ''
            if bon.plate_delivered_at and machine_off_dt:
                time_delta = bon.plate_delivered_at - machine_off_dt
                total_downtime_hours = round(time_delta.total_seconds() / 3600, 2) # Dibulatkan 2 desimal
            
            row_data = [
                bon.tanggal.strftime('%Y-%m-%d') if bon.tanggal else '',
                str(bon.pic or ''),
                str(bon.mesin_cetak or ''),
                str(bon.remarks or ''),
                str(bon.wo_number or ''),
                str(bon.mc_number or ''),
                str(bon.run_length or ''),
                str(bon.item_name or ''),
                str(bon.jumlah_plate or ''),
                str(bon.note or ''),
                str(bon.machine_off_at or ''),
                bon.plate_start_at.strftime('%Y-%m-%d %H:%M:%S') if bon.plate_start_at else '',
                bon.plate_finish_at.strftime('%Y-%m-%d %H:%M:%S') if bon.plate_finish_at else '',
                bon.plate_delivered_at.strftime('%Y-%m-%d %H:%M:%S') if bon.plate_delivered_at else '',
                str(bon.ctp_by or ''),
                str(bon.ctp_group or ''),
                str(bon.status or ''),
                total_downtime_hours
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=bon_press_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exporting bon press data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
@export_bp.route('/export-kpi-ctp', methods=['GET'])
def export_kpi_ctp():
    try:

        # Get filter parameters
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        ctp_group = request.args.get('ctp_group', '')

        # Build query for KPI CTP data
        query = CTPProductionLog.query

        # Apply filters
        if date_from:
            query = query.filter(CTPProductionLog.log_date >= date_from)
        if date_to:
            query = query.filter(CTPProductionLog.log_date <= date_to)
        if ctp_group:
            query = query.filter(CTPProductionLog.ctp_group == ctp_group)

        # Get CTP data sorted by date (newest first)
        kpi_data = query.order_by(CTPProductionLog.log_date.desc(), CTPProductionLog.id.desc()).all()

        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "KPI CTP Data"

        # Write headers
        headers = [
            'Tanggal', 'Group CTP', 'Shift', 'PIC', 'Mesin CTP',
            'Processor Temperature', 'Dwell Time', 'WO Number', 'MC Number',
            'Run Length', 'Print Machine', 'Remarks Job', 'Item Name',
            'Note', 'Plate Type Material', 'Paper Type', 'Raster',
            'Plate Good', 'Plate Not Good', 'Not Good Reason', 'Calibration'
            'Cyan 20%', 'Cyan 25%', 'Cyan 40%', 'Cyan 50%', 'Cyan 75%', 'Cyan 80%',
            'Magenta 20%', 'Magenta 25%', 'Magenta 40%', 'Magenta 50%', 'Magenta 75%', 'Magenta 80%',
            'Yellow 20%', 'Yellow 25%', 'Yellow 40%', 'Yellow 50%', 'Yellow 75%', 'Yellow 80%',
            'Black 20%', 'Black 25%', 'Black 40%', 'Black 50%', 'Black 75%', 'Black 80%',
            'Spot Color 1 20%', 'Spot Color 1 25%', 'Spot Color 1 40%', 'Spot Color 1 50%', 'Spot Color 1 75%', 'Spot Color 1 80%',
            'Spot Color 2 20%', 'Spot Color 2 25%', 'Spot Color 2 40%', 'Spot Color 2 50%', 'Spot Color 2 75%', 'Spot Color 2 80%',
            'Spot Color 3 20%', 'Spot Color 3 25%', 'Spot Color 3 40%', 'Spot Color 3 50%', 'Spot Color 3 75%', 'Spot Color 3 80%',
            'Spot Color 4 20%', 'Spot Color 4 25%', 'Spot Color 4 40%', 'Spot Color 4 50%', 'Spot Color 4 75%', 'Spot Color 4 80%',
            'Spot Color 5 20%', 'Spot Color 5 25%', 'Spot Color 5 40%', 'Spot Color 5 50%', 'Spot Color 5 75%', 'Spot Color 5 80%',
            'Spot Color 6 20%', 'Spot Color 6 25%', 'Spot Color 6 40%', 'Spot Color 6 50%', 'Spot Color 6 75%', 'Spot Color 6 80%',
            'Spot Color 7 20%', 'Spot Color 7 25%', 'Spot Color 7 40%', 'Spot Color 7 50%', 'Spot Color 7 75%', 'Spot Color 7 80%',
            'Spot Color 8 20%', 'Spot Color 8 25%', 'Spot Color 8 40%', 'Spot Color 8 50%', 'Spot Color 8 75%', 'Spot Color 8 80%',
            'Start Time', 'Finish Time'
        ]
        worksheet.append(headers)

        # Write data rows
        for kpi in kpi_data:
            row_data = [
                kpi.log_date.strftime('%Y-%m-%d') if kpi.log_date else '',
                str(kpi.ctp_group or ''),
                str(kpi.ctp_shift or ''),
                str(kpi.ctp_pic or ''),
                str(kpi.ctp_machine or ''),
                str(kpi.processor_temperature or ''),
                str(kpi.dwell_time or ''),
                str(kpi.wo_number or ''),
                str(kpi.mc_number or ''),
                str(kpi.run_length_sheet or ''),
                str(kpi.print_machine or ''),
                str(kpi.remarks_job or ''),
                str(kpi.item_name or ''),
                str(kpi.note or ''),
                str(kpi.plate_type_material or ''),
                str(kpi.paper_type or ''),
                str(kpi.raster or ''),
                str(kpi.num_plate_good or ''),
                str(kpi.num_plate_not_good or ''),
                str(kpi.not_good_reason or ''),
                str(kpi.calibration or ''),
                str(kpi.cyan_20_percent or ''),
                str(kpi.cyan_25_percent or ''),
                str(kpi.cyan_40_percent or ''),                
                str(kpi.cyan_50_percent or ''),
                str(kpi.cyan_75_percent or ''),
                str(kpi.cyan_80_percent or ''),
                str(kpi.magenta_20_percent or ''),
                str(kpi.magenta_25_percent or ''),
                str(kpi.magenta_40_percent or ''),                
                str(kpi.magenta_50_percent or ''),
                str(kpi.magenta_75_percent or ''),
                str(kpi.magenta_80_percent or ''),
                str(kpi.yellow_20_percent or ''),
                str(kpi.yellow_25_percent or ''),
                str(kpi.yellow_40_percent or ''),                
                str(kpi.yellow_50_percent or ''),
                str(kpi.yellow_75_percent or ''),
                str(kpi.yellow_80_percent or ''),
                str(kpi.black_20_percent or ''),
                str(kpi.black_25_percent or ''),
                str(kpi.black_40_percent or ''),                
                str(kpi.black_50_percent or ''),
                str(kpi.black_75_percent or ''),
                str(kpi.black_80_percent or ''),
                str(kpi.x_20_percent or ''),
                str(kpi.x_25_percent or ''),
                str(kpi.x_40_percent or ''),                
                str(kpi.x_50_percent or ''),
                str(kpi.x_75_percent or ''),
                str(kpi.x_80_percent or ''),
                str(kpi.z_20_percent or ''),
                str(kpi.z_25_percent or ''),
                str(kpi.z_40_percent or ''),                
                str(kpi.z_50_percent or ''),
                str(kpi.z_75_percent or ''),
                str(kpi.z_80_percent or ''),
                str(kpi.u_20_percent or ''),
                str(kpi.u_25_percent or ''),
                str(kpi.u_40_percent or ''),
                str(kpi.u_50_percent or ''),
                str(kpi.u_80_percent or ''),
                str(kpi.u_75_percent or ''),
                str(kpi.v_20_percent or ''),
                str(kpi.v_25_percent or ''),
                str(kpi.v_40_percent or ''),
                str(kpi.v_50_percent or ''),
                str(kpi.v_80_percent or ''),
                str(kpi.v_75_percent or ''),
                str(kpi.f_20_percent or ''),
                str(kpi.f_25_percent or ''),
                str(kpi.f_40_percent or ''),
                str(kpi.f_50_percent or ''),
                str(kpi.f_80_percent or ''),
                str(kpi.f_75_percent or ''),
                str(kpi.g_20_percent or ''),
                str(kpi.g_25_percent or ''),
                str(kpi.g_40_percent or ''),
                str(kpi.g_50_percent or ''),
                str(kpi.g_80_percent or ''),
                str(kpi.g_75_percent or ''),
                str(kpi.h_20_percent or ''),
                str(kpi.h_25_percent or ''),
                str(kpi.h_40_percent or ''),
                str(kpi.h_50_percent or ''),
                str(kpi.h_80_percent or ''),
                str(kpi.h_75_percent or ''),
                str(kpi.j_20_percent or ''),
                str(kpi.j_25_percent or ''),
                str(kpi.j_40_percent or ''),
                str(kpi.j_50_percent or ''),
                str(kpi.j_80_percent or ''),
                str(kpi.j_75_percent or ''),
                kpi.start_time.strftime('%H:%M') if kpi.start_time else '',
                kpi.finish_time.strftime('%H:%M') if kpi.finish_time else ''
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=kpi_ctp_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
    
    except Exception as e:
        print(f"Error exporting KPI CTP data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@export_bp.route('/export-ctp-bon', methods=['GET'])
def export_ctp_bon():
    try:
        
        # Get filter parameters
        status_filter = request.args.get('status', '')
        remarks_filter = request.args.get('remarks', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Build query
        query = PlateBonRequest.query
        
        # Apply filters
        if status_filter:
            query = query.filter(PlateBonRequest.status == status_filter)
        
        if remarks_filter:
            query = query.filter(PlateBonRequest.remarks.ilike(f'%{remarks_filter}%'))
            
        if date_from:
            query = query.filter(PlateBonRequest.tanggal >= date_from)
        if date_to:
            query = query.filter(PlateBonRequest.tanggal <= date_to)
        
        # Get data sorted by date (newest first)
        bon_requests = query.order_by(PlateBonRequest.tanggal.desc(), PlateBonRequest.id.desc()).all()
        
        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "CTP Bon Data"
        
        # Write headers
        headers = [
            'ID', 'Tanggal', 'Mesin Cetak', 'PIC', 'Remarks', 'WO Number', 'MC Number', 
            'Run Length', 'Item Name', 'Jumlah Plate', 'Note', 'Machine Off At',
            'Plate Start At', 'Plate Finish At', 'Plate Delivered At', 'CTP By', 'Status'
        ]
        worksheet.append(headers)
        
        # Write data rows
        for bon in bon_requests:
            row_data = [
                str(bon.id or ''),
                bon.tanggal.strftime('%Y-%m-%d') if bon.tanggal else '',
                str(bon.mesin_cetak or ''),
                str(bon.pic or ''),
                str(bon.remarks or ''),
                str(bon.wo_number or ''),
                str(bon.mc_number or ''),
                str(bon.run_length or ''),
                str(bon.item_name or ''),
                str(bon.jumlah_plate or ''),
                str(bon.note or ''),
                bon.machine_off_at.strftime('%Y-%m-%d %H:%M:%S') if bon.machine_off_at else '',
                bon.plate_start_at.strftime('%Y-%m-%d %H:%M:%S') if bon.plate_start_at else '',
                bon.plate_finish_at.strftime('%Y-%m-%d %H:%M:%S') if bon.plate_finish_at else '',
                bon.plate_delivered_at.strftime('%Y-%m-%d %H:%M:%S') if bon.plate_delivered_at else '',
                str(bon.ctp_by or ''),
                str(bon.status or '')
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=ctp_bon_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exporting CTP bon data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@export_bp.route('/export-ctp-bon-data', methods=['GET'])
def export_ctp_bon_data():
    try:
        
        # Get filter parameters
        status_filter = request.args.get('status', '')
        remarks_filter = request.args.get('remarks', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Define the filename with .xlsx extension
        filename = f'ctp_bon_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Build query
        query = PlateBonRequest.query
        
        # Apply filters
        if status_filter:
            if status_filter == 'selesai':
                query = query.filter(PlateBonRequest.status == 'selesai')
            else:
                query = query.filter(PlateBonRequest.status == status_filter)
        
        if remarks_filter:
            query = query.filter(PlateBonRequest.remarks.ilike(f'%{remarks_filter}%'))
            
        if start_date:
            query = query.filter(PlateBonRequest.tanggal >= start_date)
        if end_date:
            query = query.filter(PlateBonRequest.tanggal <= end_date)
        
        # Get data sorted by date (newest first)
        bon_requests = query.order_by(PlateBonRequest.tanggal.desc(), PlateBonRequest.id.desc()).all()
        
        # Create an in-memory Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "CTP Bon Data"
        
        # Write headers
        headers = [
            'ID', 'Tanggal', 'Mesin Cetak', 'PIC', 'Remarks', 'WO Number', 'MC Number', 
            'Run Length', 'Item Name', 'Jumlah Plate', 'Note', 'Machine Off At',
            'Plate Start At', 'Plate Finish At', 'Plate Delivered At', 'CTP By', 'Status'
        ]
        worksheet.append(headers)
        
        # Write data rows
        for bon in bon_requests:
            row_data = [
                str(bon.id or ''),
                bon.tanggal.strftime('%Y-%m-%d') if bon.tanggal else '',
                str(bon.mesin_cetak or ''),
                str(bon.pic or ''),
                str(bon.remarks or ''),
                str(bon.wo_number or ''),
                str(bon.mc_number or ''),
                str(bon.run_length or ''),
                str(bon.item_name or ''),
                str(bon.jumlah_plate or ''),
                str(bon.note or ''),
                bon.machine_off_at.strftime('%Y-%m-%d %H:%M:%S') if bon.machine_off_at else '',
                bon.plate_start_at.strftime('%Y-%m-%d %H:%M:%S') if bon.plate_start_at else '',
                bon.plate_finish_at.strftime('%Y-%m-%d %H:%M:%S') if bon.plate_finish_at else '',
                bon.plate_delivered_at.strftime('%Y-%m-%d %H:%M:%S') if bon.plate_delivered_at else '',
                str(bon.ctp_by or ''),
                str(bon.status or '')
            ]
            worksheet.append(row_data)
        
        # Create response with proper Excel headers
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        print(f"Error exporting CTP bon data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500